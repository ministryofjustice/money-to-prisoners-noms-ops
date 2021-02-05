import base64
import copy
from contextlib import contextmanager
import datetime
import itertools
import json
import logging
import re
import tempfile
from unittest import mock
from urllib.parse import parse_qs, urlencode

from django.conf import settings
from django.core import mail
from django.http import QueryDict
from django.test import SimpleTestCase, override_settings
from django.urls import reverse
from django.utils.dateparse import parse_datetime
from django.utils.timezone import make_aware
from mtp_common.auth import USER_DATA_SESSION_KEY, urljoin
from mtp_common.auth.api_client import get_request_token_url
from mtp_common.auth.test_utils import generate_tokens
from mtp_common.test_utils import silence_logger
from openpyxl import load_workbook
from parameterized import parameterized
import pytz
import responses

from security import (
    confirmed_prisons_flag,
    hmpps_employee_flag,
    not_hmpps_employee_flag,
    required_permissions,
    provided_job_info_flag,
)
from security.constants import (
    CHECK_REJECTION_CATEGORY_TEXT_MAPPING,
    CHECK_REJECTION_CATEGORY_BOOLEAN_MAPPING,
    CHECK_AUTO_ACCEPT_UNIQUE_CONSTRAINT_ERROR
)
from security.forms.object_list import PrisonSelectorSearchFormMixin, PRISON_SELECTOR_USER_PRISONS_CHOICE_VALUE
from security.models import EmailNotifications
from security.tests import api_url, TEST_IMAGE_DATA
from security.views.check import AcceptOrRejectCheckView
from security.views.object_base import SEARCH_FORM_SUBMITTED_INPUT_NAME

SAMPLE_PRISONS = [
    {
        'nomis_id': 'AAI',
        'general_ledger_code': '001',
        'name': 'HMP & YOI Test 1',
        'short_name': 'Test 1',
        'region': 'London',
        'categories': [{'description': 'Category D', 'name': 'D'},
                       {'description': 'Young Offender Institution', 'name': 'YOI'}],
        'populations': [{'description': 'Female', 'name': 'female'},
                        {'description': 'Male', 'name': 'male'},
                        {'description': 'Young offenders', 'name': 'young'}],
        'pre_approval_required': False,
    },
    {
        'nomis_id': 'BBI',
        'general_ledger_code': '002',
        'name': 'HMP Test 2',
        'short_name': 'Test 2',
        'region': 'London',
        'categories': [{'description': 'Category D', 'name': 'D'}],
        'populations': [{'description': 'Male', 'name': 'male'}],
        'pre_approval_required': False,
    },
]


def mock_prison_response(rsps=None):
    rsps = rsps or responses
    rsps.add(
        rsps.GET,
        api_url('/prisons/'),
        json={
            'count': len(SAMPLE_PRISONS),
            'results': SAMPLE_PRISONS,
        }
    )


def no_saved_searches(rsps=None):
    rsps = rsps or responses
    rsps.add(
        rsps.GET,
        api_url('/searches/'),
        json={
            'count': 0,
            'results': []
        },
    )


def mock_post_for_job_info_endpoint(rsps=None):
    rsps = rsps or responses
    rsps.add(
        rsps.POST,
        api_url('/job-information/'),
        json={
            'job_title': 'Evidence collator',
            'prison_estate': 'National',
            'tasks': 'Inmate Processing'
        },
    )


def offset_isodatetime_by_ten_seconds(isodatetime, offset_multiplier=1):
    return (
        parse_datetime(isodatetime)
        + datetime.timedelta(seconds=offset_multiplier * 10)
    ).isoformat()


def generate_rejection_category_test_cases_text():
    return [
        (text_category, f'i_am_a_{text_category}_value')
        for text_category in CHECK_REJECTION_CATEGORY_TEXT_MAPPING.keys()
    ]


def generate_rejection_category_test_cases_bool():
    return [
        (text_category, True)
        for text_category in CHECK_REJECTION_CATEGORY_BOOLEAN_MAPPING.keys()
    ]


@override_settings(
    CACHES={'default': {'BACKEND': 'django.core.cache.backends.dummy.DummyCache'}},
)
class SecurityBaseTestCase(SimpleTestCase):

    def setUp(self):
        super().setUp()
        self.notifications_mock = mock.patch('mtp_common.templatetags.mtp_common.notifications_for_request',
                                             return_value=[])
        self.notifications_mock.start()

    def tearDown(self):
        self.notifications_mock.stop()
        super().tearDown()

    def login(self, rsps, follow=True, user_data=None):
        no_saved_searches(rsps=rsps)
        return self._login(rsps, follow=follow, user_data=user_data)

    def login_test_searches(self, rsps, follow=True):
        return self._login(rsps, follow=follow)

    def _login(self, rsps, follow=True, user_data=None):
        returned_user_data = user_data or self.get_user_data()
        returned_user_data['pk'] = 5
        rsps.add(
            rsps.POST,
            get_request_token_url(),
            json=generate_tokens()
        )
        rsps.add(
            rsps.GET,
            urljoin(settings.API_URL, '/users/{username}'.format(username=returned_user_data['username'])),
            json=returned_user_data
        )

        response = self.client.post(
            reverse('login'),
            data={'username': 'shall', 'password': 'pass'},
            follow=follow
        )
        if follow:
            self.assertEqual(response.status_code, 200)
        else:
            self.assertEqual(response.status_code, 302)
        return response

    def get_user_data(
        self,
        first_name='Sam',
        last_name='Hall',
        username='shall',
        email='sam@mtp.local',
        permissions=required_permissions,
        prisons=(SAMPLE_PRISONS[1],),
        flags=(hmpps_employee_flag, confirmed_prisons_flag, provided_job_info_flag),
        roles=('security',),
    ):
        return {
            'first_name': first_name,
            'last_name': last_name,
            'username': username,
            'email': email,
            'permissions': permissions,
            'prisons': prisons,
            'flags': flags,
            'roles': roles,
        }

    def assertSpreadsheetEqual(self, spreadsheet_data, expected_values, msg=None):  # noqa: N802
        with temp_spreadsheet(spreadsheet_data) as ws:
            for i, row in enumerate(expected_values, start=1):
                for j, cell in enumerate(row, start=1):
                    self.assertEqual(cell, ws.cell(column=j, row=i).value, msg=msg)


class LocaleTestCase(SecurityBaseTestCase):
    def test_locale_switches_based_on_browser_language(self):
        languages = (
            ('*', 'en-gb'),
            ('en', 'en-gb'),
            ('en-gb', 'en-gb'),
            ('en-GB, en, *', 'en-gb'),
            ('cy', 'cy'),
            ('cy, en-GB, en, *', 'cy'),
            ('en, cy, *', 'en-gb'),
            ('es', 'en-gb'),
        )
        with silence_logger(name='django.request', level=logging.ERROR):
            for accept_language, expected_slug in languages:
                response = self.client.get('/', HTTP_ACCEPT_LANGUAGE=accept_language)
                self.assertRedirects(response, '/%s/' % expected_slug, fetch_redirect_response=False)
                response = self.client.get('/login/', HTTP_ACCEPT_LANGUAGE=accept_language)
                self.assertRedirects(response, '/%s/login/' % expected_slug, fetch_redirect_response=True)


class SecurityDashboardViewsTestCase(SecurityBaseTestCase):
    @responses.activate
    def test_can_access_security_dashboard(self):
        response = self.login(responses)
        self.assertContains(response, '<!-- security:dashboard -->')

    @responses.activate
    def test_cannot_access_prisoner_location_admin(self):
        self.login(responses)
        no_saved_searches()
        response = self.client.get(reverse('location_file_upload'), follow=True)
        self.assertNotContains(response, '<!-- location_file_upload -->')
        self.assertContains(response, '<!-- security:dashboard -->')


class PrisonSwitcherTestCase(SecurityBaseTestCase):
    """
    Tests related to the prison switcher area on the top of some pages.
    The prison switcher area shows the prisons that the user selected in the settings
    page with a link to change this.
    """

    def _mock_api_responses(self):
        no_saved_searches()
        mock_prison_response()
        responses.add(
            responses.GET,
            api_url('/senders/'),
            json={},
        )

    @responses.activate
    def test_with_many_prisons(self):
        """
        Test that if the user has more than 4 prisons in settings, only the first ones are
        shown in the prison switcher area to avoid long text.
        """
        self._mock_api_responses()
        prisons = [
            {
                **SAMPLE_PRISONS[0],
                'name': f'Prison {index}',
            } for index in range(1, 11)
        ]
        self.login(
            responses,
            user_data=self.get_user_data(prisons=prisons),
        )
        response = self.client.get(reverse('security:sender_list'))
        self.assertContains(
            response,
            'Prison 1, Prison 2, Prison 3, Prison 4',
        )
        self.assertContains(
            response,
            ' and 6 more',
        )

    @responses.activate
    def test_with_fewer_prisons(self):
        """
        Test that if the user has less than 4 prisons in settings,
        they are all shown in the prison switcher area.
        """
        self._mock_api_responses()
        prisons = [
            {
                **SAMPLE_PRISONS[0],
                'name': f'Prison {index}',
            } for index in range(1, 3)
        ]
        self.login(
            responses,
            user_data=self.get_user_data(prisons=prisons),
        )
        response = self.client.get(reverse('security:sender_list'))
        self.assertContains(
            response,
            'Prison 1, Prison 2',
        )

        self.assertNotContains(
            response,
            'Prison 3',
        )

    @responses.activate
    def test_sees_all_prisons(self):
        """
        Test that if the user hasn't specified any prisons in settings, it means that he/she can
        see all prisons so the text 'All prisons' is known in the prison switcher area.
        """
        self._mock_api_responses()
        self.login(
            responses,
            user_data=self.get_user_data(prisons=[]),
        )
        response = self.client.get(reverse('security:sender_list'))
        self.assertContains(
            response,
            'All prisons',
        )


class JobInformationViewTestCase(SecurityBaseTestCase):
    protected_views = [
        'security:credit_list',
        'security:dashboard',
        'security:disbursement_list',
        'security:prisoner_list',
        'security:sender_list',
    ]

    @responses.activate
    def test_redirects_when_provided_job_info_flag_is_missing(self):
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[
                    confirmed_prisons_flag,
                    hmpps_employee_flag,
                ],
            ),
        )
        for view in self.protected_views:
            response = self.client.get(reverse(view), follow=True)
            self.assertContains(response, '<!-- job_information -->')

    @responses.activate
    def test_can_view_app_when_user_has_provided_job_info_flag(self):
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[
                    confirmed_prisons_flag,
                    hmpps_employee_flag,
                    provided_job_info_flag,
                ]
            ),
        )
        response = self.client.get(reverse('security:dashboard'), follow=True)
        self.assertContains(response, '<!-- security:dashboard -->')

    @responses.activate
    def test_successful_form_post(self):
        mock_post_for_job_info_endpoint()
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[confirmed_prisons_flag, hmpps_employee_flag]
            )
        )
        responses.add(
            responses.PUT,
            api_url('/users/shall/flags/%s/' % provided_job_info_flag),
            json={}
        )

        response = self.client.post(reverse('job_information'), data={
            'job_title': 'Evidence collator',
            'prison_estate': 'National',
            'tasks': 'yas sir'
        }, follow=True)

        self.assertContains(response, '<!-- security:dashboard -->')

    @responses.activate
    def test_unsuccessful_form_post(self):
        mock_post_for_job_info_endpoint()
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[
                    confirmed_prisons_flag,
                    hmpps_employee_flag
                ]
            )
        )
        responses.add(
            responses.PUT,
            api_url('/users/shall/flags/%s/' % provided_job_info_flag),
            json={}
        )

        response = self.client.post(reverse('job_information'), data={
            'job_title': '',
            'prison_estate': '',
            'tasks': ''
        }, follow=True)

        self.assertContains(response, '<!-- job_information -->')
        self.assertContains(response, 'This field is required')

    @responses.activate
    def test_form_adds_provided_job_info_flag(self):
        mock_post_for_job_info_endpoint()
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[
                    confirmed_prisons_flag,
                    hmpps_employee_flag
                ]
            )
        )
        responses.add(
            responses.PUT,
            api_url('/users/shall/flags/%s/' % provided_job_info_flag),
            json={}
        )

        response = self.client.post(reverse('job_information'), data={
            'job_title': 'Evidence collator',
            'prison_estate': 'National',
            'tasks': 'yas sir'
        }, follow=True)

        self.assertIn(provided_job_info_flag, response.context['user'].user_data['flags'])


class HMPPSEmployeeTestCase(SecurityBaseTestCase):
    protected_views = [
        'security:credit_list',
        'security:dashboard',
        'security:disbursement_list',
        'security:prisoner_list',
        'security:sender_list',
    ]

    @responses.activate
    def test_redirects_when_no_flag(self):
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[
                    confirmed_prisons_flag,
                ],
            ),
        )
        for view in self.protected_views:
            response = self.client.get(reverse(view), follow=True)
            self.assertContains(response, '<!-- security:hmpps_employee -->')

    @responses.activate
    def test_non_employee_flag_disallows_entry(self):
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[
                    confirmed_prisons_flag,
                    not_hmpps_employee_flag,
                ],
            ),
        )
        for view in self.protected_views:
            response = self.client.get(reverse(view), follow=True)
            self.assertContains(response, '<!-- security:not_hmpps_employee -->')
            self.assertIn('You can’t use this tool', response.content.decode())

    @responses.activate
    def test_employee_can_access(self):
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[
                    confirmed_prisons_flag,
                    hmpps_employee_flag,
                    provided_job_info_flag,
                ]
            )
        )

        def assertViewAccessible(view):  # noqa: N802
            response = self.client.get(reverse(view), follow=True)
            self.assertContains(response, '<!-- %s -->' % view)

        assertViewAccessible('security:dashboard')
        mock_prison_response()
        assertViewAccessible('security:credit_list')

    @responses.activate
    def test_employee_flag_set(self):
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=['abc', confirmed_prisons_flag, provided_job_info_flag]
            )
        )
        responses.add(
            responses.PUT,
            api_url('/users/shall/flags/%s/' % hmpps_employee_flag),
            json={}
        )
        response = self.client.post(reverse('security:hmpps_employee'), data={
            'confirmation': 'yes',
        }, follow=True)
        self.assertContains(response, '<!-- security:dashboard -->')
        self.assertIn(hmpps_employee_flag, self.client.session[USER_DATA_SESSION_KEY]['flags'])
        self.assertIn(hmpps_employee_flag, response.context['user'].user_data['flags'])

    @responses.activate
    def test_redirects_to_referrer(self):
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=[
                    confirmed_prisons_flag,
                    provided_job_info_flag,
                ],
            ),
        )
        responses.add(
            responses.PUT,
            api_url('/users/shall/flags/%s/' % hmpps_employee_flag),
            json={}
        )
        mock_prison_response()
        response = self.client.post(reverse('security:hmpps_employee'), data={
            'confirmation': 'yes',
            'next': reverse('security:prisoner_list'),
        }, follow=True)
        self.assertContains(response, '<!-- security:prisoner_list -->')

    @responses.activate
    def test_non_employee_flag_set(self):
        self.login(
            responses,
            user_data=self.get_user_data(
                flags=['123', confirmed_prisons_flag]
            )
        )
        responses.add(
            responses.PUT,
            api_url('/users/shall/flags/%s/' % not_hmpps_employee_flag),
            json={}
        )
        responses.add(
            responses.DELETE,
            api_url('/users/shall/'),
            json={}
        )
        response = self.client.post(reverse('security:hmpps_employee'), data={
            'confirmation': 'no',
        }, follow=True)
        self.assertContains(response, '<!-- security:not_hmpps_employee -->')
        self.assertIn('You can’t use this tool', response.content.decode())


class SecurityViewTestCase(SecurityBaseTestCase):
    view_name = None
    api_list_path = None
    search_ordering = None

    bank_transfer_sender = {
        'id': 9,
        'credit_count': 4,
        'credit_total': 41000,
        'prisoner_count': 3,
        'prison_count': 2,
        'prisons': [
            {
                'nomis_id': 'PRN',
                'name': 'Prison PRN',
            },
            {
                'nomis_id': 'ABC',
                'name': 'Prison ABC',
            },
        ],
        'bank_transfer_details': [
            {
                'sender_name': 'MAISIE NOLAN',
                'sender_sort_code': '101010',
                'sender_account_number': '12312345',
                'sender_roll_number': '',
            }
        ],
        'created': '2016-05-25T20:24:00Z',
        'modified': '2016-05-25T20:24:00Z',
    }
    debit_card_sender = {
        'id': 9,
        'credit_count': 4,
        'credit_total': 42000,
        'prisoner_count': 3,
        'prison_count': 2,
        'prisons': [
            {
                'nomis_id': 'PRN',
                'name': 'Prison PRN',
            },
            {
                'nomis_id': 'ABC',
                'name': 'Prison ABC',
            },
        ],
        'bank_transfer_details': [],
        'debit_card_details': [
            {
                'card_number_last_digits': '1234',
                'card_expiry_date': '10/20',
                'postcode': 'SW137NJ',
                'sender_emails': ['m@outside.local', 'M@OUTSIDE.LOCAL', 'mn@outside.local'],
                'cardholder_names': ['Maisie N', 'MAISIE N', 'Maisie Nolan'],
            }
        ],
        'created': '2016-05-25T20:24:00Z',
        'modified': '2016-05-25T20:24:00Z',
    }
    prisoner_profile = {
        'id': 8,
        'credit_count': 3,
        'credit_total': 31000,
        'sender_count': 2,
        'disbursement_count': 2,
        'disbursement_total': 29000,
        'recipient_count': 1,
        'prisoner_name': 'JAMES HALLS',
        'prisoner_number': 'A1409AE',
        'prisoner_dob': '1986-12-09',
        'current_prison': {'nomis_id': 'PRN', 'name': 'Prison PRN'},
        'prisons': [{'nomis_id': 'PRN', 'name': 'Prison PRN'}],
        'provided_names': ['Jim Halls', 'JAMES HALLS', 'James Halls '],
        'created': '2016-05-25T20:24:00Z',
        'modified': '2016-05-25T20:24:00Z',
    }
    credit_object = {
        'id': 1,
        'amount': 10250,
        'resolution': 'credited', 'anonymous': False,
        'credited_at': '2016-05-25T20:24:00Z',
        'owner': 1, 'owner_name': 'Clerk',
        'prison': 'PRN', 'prison_name': 'Prison', 'prisoner_name': 'JAMES HALLS', 'prisoner_number': 'A1409AE',
        'reconciliation_code': None,
        'received_at': '2017-01-25T12:00:00Z',
        'refunded_at': None,
        'reviewed': False, 'comments': [],
        'source': 'bank_transfer',
        'sender_sort_code': '101010', 'sender_account_number': '12312345', 'sender_roll_number': '',
        'card_expiry_date': None, 'card_number_last_digits': None,
        'sender_name': 'MAISIE NOLAN',
        'sender_email': None,
        'intended_recipient': None,
    }

    def get_api_object_list_response_data(self):
        """
        Return the list of objects that the mocked api should return in the tests.
        """
        return []

    @responses.activate
    @mock.patch('security.forms.object_base.SecurityForm.get_object_list')
    def test_can_access_security_view(self, mocked_form_method):
        mocked_form_method.return_value = []
        if not self.view_name:
            return
        mock_prison_response()
        self.login(responses)
        response = self.client.get(reverse(self.view_name), follow=True)
        self.assertContains(response, '<!-- %s -->' % self.view_name)


class SearchV2SecurityTestCaseMixin:
    search_results_view_name = None
    advanced_search_view_name = None

    # the filter name used for API calls, it's usually prison but can sometimes be current_prison
    prison_api_filter_name = 'prison'

    def test_displays_simple_search_results(self):
        """
        Test that the search results page includes the objects returned by the API
        in case of a simple search.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            api_results = self.get_api_object_list_response_data()
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': len(api_results),
                    'results': api_results,
                },
            )
            response = self.client.get(reverse(self.search_results_view_name))
        self._test_search_results_content(response, advanced=False)

    def test_displays_advanced_search_results(self):
        """
        Test that the search results page includes the objects returned by the API
        in case of an advanced search.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            api_results = self.get_api_object_list_response_data()
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': len(api_results),
                    'results': api_results,
                },
            )
            response = self.client.get(
                f'{reverse(self.search_results_view_name)}?advanced=True'
            )
        self._test_search_results_content(response, advanced=True)

    def _test_search_results_content(self, response, advanced=False):
        """
        Subclass to test that the response content of the search results view is as expected.
        """
        raise NotImplementedError

    def test_simple_search_uses_default_prisons(self):
        """
        Test that the simple search template uses the prisons of the logged in user to
        populate the `prison` hidden inputs.
        The simple search filters by the user's prisons by default.
        """
        with responses.RequestsMock() as rsps:
            user_data = self.get_user_data(prisons=SAMPLE_PRISONS)

            self.login(rsps, user_data=user_data)
            mock_prison_response(rsps=rsps)
            response = self.client.get(reverse(self.view_name))

        self.assertContains(
            response,
            f'<input type="hidden" name="prison_selector" value="{PRISON_SELECTOR_USER_PRISONS_CHOICE_VALUE}"',
        )

    def test_advanced_search_with_my_prisons_selection(self):
        """
        Test that if prison_selector == mine, the API call includes the prison filter with
        the expanted current user's prisons value.
        """
        with responses.RequestsMock() as rsps:
            user_prisons = SAMPLE_PRISONS[:1]
            user_data = self.get_user_data(prisons=user_prisons)

            self.login(rsps, user_data=user_data)
            mock_prison_response(rsps=rsps)

            url = (
                f'{reverse(self.advanced_search_view_name)}'
                f'?prison_selector={PRISON_SELECTOR_USER_PRISONS_CHOICE_VALUE}'
                f'&advanced=True&{SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            )
            response = self.client.get(url, follow=True)
            self.assertEqual(response.status_code, 200)

            api_call_made = rsps.calls[-1].request.url
            parsed_qs = parse_qs(api_call_made.split('?', 1)[1])
            self.assertCountEqual(
                parsed_qs[self.prison_api_filter_name],
                [prison['nomis_id'] for prison in user_prisons],
            )

    def test_advanced_search_with_all_prisons_selection(self):
        """
        Test that if prison_selector == all, the API call doesn't include any prison filter.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)

            url = (
                f'{reverse(self.advanced_search_view_name)}'
                f'?prison_selector={PrisonSelectorSearchFormMixin.PRISON_SELECTOR_ALL_PRISONS_CHOICE_VALUE}'
                f'&advanced=True&{SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            )
            response = self.client.get(url, follow=True)
            self.assertEqual(response.status_code, 200)

            api_call_made = rsps.calls[-1].request.url
            parsed_qs = parse_qs(api_call_made.split('?', 1)[1])
            self.assertTrue(self.prison_api_filter_name not in parsed_qs)

    def test_advanced_search_with_exact_prison_selected(self):
        """
        Test that if prison_selector == exact and a prison is specified, the API call
        includes exactly that prison filter.
        """
        with responses.RequestsMock() as rsps:
            expected_prison_id = SAMPLE_PRISONS[1]['nomis_id']

            self.login(rsps)
            mock_prison_response(rsps=rsps)

            url = (
                f'{reverse(self.advanced_search_view_name)}'
                f'?prison_selector={PrisonSelectorSearchFormMixin.PRISON_SELECTOR_EXACT_PRISON_CHOICE_VALUE}'
                f'&prison={expected_prison_id}'
                f'&advanced=True&{SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            )
            response = self.client.get(url, follow=True)
            self.assertEqual(response.status_code, 200)

            api_call_made = rsps.calls[-1].request.url
            parsed_qs = parse_qs(api_call_made.split('?', 1)[1])
            self.assertCountEqual(
                parsed_qs[self.prison_api_filter_name],
                [expected_prison_id],
            )

    def test_simple_search_redirects_to_search_results_page(self):
        """
        Test that submitting the simple search form redirects to the results page when the form is valid.
        The action of submitting the form is represented by the query param
        SEARCH_FORM_SUBMITTED_INPUT_NAME which gets removed when redirecting.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )
            query_string = (
                f'prison_selector={PRISON_SELECTOR_USER_PRISONS_CHOICE_VALUE}'
                f'&advanced=False&ordering={self.search_ordering}&simple_search=test'
            )
            request_url = f'{reverse(self.view_name)}?{query_string}&{SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            expected_redirect_url = f'{reverse(self.search_results_view_name)}?{query_string}'
            response = self.client.get(request_url)
            self.assertRedirects(
                response,
                expected_redirect_url,
            )

    def test_advanced_search_redirects_to_search_results_page(self):
        """
        Test that submitting the advanced search form redirects to the results page when the form is valid.
        The action of submitting the form is represented by the query param
        SEARCH_FORM_SUBMITTED_INPUT_NAME which gets removed when redirecting.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )
            query_string = (
                f'prison_selector={PRISON_SELECTOR_USER_PRISONS_CHOICE_VALUE}'
                f'&advanced=True&ordering={self.search_ordering}'
            )
            request_url = (
                f'{reverse(self.advanced_search_view_name)}?{query_string}&{SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            )
            expected_redirect_url = f'{reverse(self.search_results_view_name)}?{query_string}'
            response = self.client.get(request_url)
            self.assertRedirects(
                response,
                expected_redirect_url,
            )

    def test_simple_search_doesnt_redirect_to_search_results_page_if_form_is_invalid(self):
        """
        Test that submitting the simple search form doesn't redirect to the results page when the form is invalid.
        The action of submitting the form is represented by the query param SEARCH_FORM_SUBMITTED_INPUT_NAME.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            query_string = 'ordering=invalid&simple_search=test'
            request_url = f'{reverse(self.view_name)}?{query_string}&{SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            response = self.client.get(request_url)
            self.assertFalse(response.context['form'].is_valid())
            self.assertEqual(response.status_code, 200)

    def test_advanced_search_doesnt_redirect_to_search_results_page_if_form_is_invalid(self):
        """
        Test that submitting the advanced search form doesn't redirect to the results page when the form is invalid.
        The action of submitting the form is represented by the query param SEARCH_FORM_SUBMITTED_INPUT_NAME.
        """
        if not self.advanced_search_view_name:
            self.skipTest(f'Advanced search not yet implemented for {self.__class__.__name__}')

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            query_string = 'ordering=invalid&advanced=True'
            request_url = (
                f'{reverse(self.advanced_search_view_name)}?{query_string}&{SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            )
            response = self.client.get(request_url)
            self.assertFalse(response.context['form'].is_valid())
            self.assertEqual(response.status_code, 200)

    def test_navigating_back_to_simple_search_form(self):
        """
        Test that going back to the simple search form doesn't redirect to the results page.
        The action of NOT submitting the form explicitly is represented by the absence of query param
        SEARCH_FORM_SUBMITTED_INPUT_NAME.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )
            query_string = f'ordering={self.search_ordering}&simple_search=test'
            request_url = f'{reverse(self.view_name)}?{query_string}'
            response = self.client.get(request_url)
            self.assertEqual(response.status_code, 200)

    def test_simple_search_results_with_link_to_all_prisons_search(self):
        """
        Test that if the current user's prisons value != 'all', after making a simple search,
        a link to search for the same thing in all prisons is shown.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )

            response = self.client.get(f'{reverse(self.search_results_view_name)}?simple_search=test')

        link_re = re.compile(
            (
                '<a href="(.*)'
                f'prison_selector={PrisonSelectorSearchFormMixin.PRISON_SELECTOR_ALL_PRISONS_CHOICE_VALUE}'
                '(.*)">'
                '(.*)See results from all prisons(.*)'
                '</a>'
            ),
            re.DOTALL,
        )
        self.assertTrue(
            link_re.search(response.content.decode('utf8')),
        )

    def test_simple_search_results_without_link_to_all_prisons_search(self):
        """
        Test that if the current user's prisons value == 'all', after making a simple search
        no link to search for the same thing in all prisons is shown as it would have the
        same result.
        """
        with responses.RequestsMock() as rsps:
            user_data = self.get_user_data(prisons=None)

            self.login(rsps, user_data=user_data)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )
            response = self.client.get(f'{reverse(self.search_results_view_name)}?simple_search=test')

        link_re = re.compile(
            (
                '<a href="(.*)'
                f'prison_selector={PrisonSelectorSearchFormMixin.PRISON_SELECTOR_ALL_PRISONS_CHOICE_VALUE}'
                '(.*)">'
                '(.*)See results from all prisons(.*)'
                '</a>'
            ),
            re.DOTALL,
        )
        self.assertFalse(
            link_re.search(response.content.decode('utf8')),
        )

    def test_displays_simple_search_results_with_all_prisons(self):
        """
        Test that when making a simple search with prison_selector == all,
        the API call doesn't include any prison filter.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            api_results = self.get_api_object_list_response_data()
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': len(api_results),
                    'results': api_results,
                },
            )
            url = (
                f'{reverse(self.search_results_view_name)}'
                '?simple_search=test'
                f'&prison_selector={PrisonSelectorSearchFormMixin.PRISON_SELECTOR_ALL_PRISONS_CHOICE_VALUE}'
            )
            response = self.client.get(url)
            self.assertEqual(response.status_code, 200)

            api_call_made = rsps.calls[-1].request.url
            parsed_qs = parse_qs(api_call_made.split('?', 1)[1])
            self.assertTrue(self.prison_api_filter_name not in parsed_qs)
            self.assertTrue('simple_search' in parsed_qs)


class ExportSecurityViewTestCaseMixin:
    export_view_name = None
    export_email_view_name = None
    export_expected_xls_headers = None
    export_expected_xls_rows = None

    def test_export_some_data(self):
        """
        Test that the export view generates a spreadsheet with only the content returned
        by the API.
        """
        expected_spreadsheet_content = [
            self.export_expected_xls_headers,
            *self.export_expected_xls_rows,
        ]

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 2,
                    'results': self.get_api_object_list_response_data(),
                }
            )
            response = self.client.get(reverse(self.export_view_name))

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertSpreadsheetEqual(response.content, expected_spreadsheet_content)

    def test_email_export_some_data(self):
        """
        Test that the view sends the expected spreadsheet via email.
        """
        expected_spreadsheet_content = [
            self.export_expected_xls_headers,
            *self.export_expected_xls_rows,
        ]

        # get realistic referer
        qs = f'ordering={self.search_ordering}'
        response = self.client.get('/')
        referer_url = response.wsgi_request.build_absolute_uri(
            f'{reverse(self.view_name)}?{qs}',
        )

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 2,
                    'results': self.get_api_object_list_response_data(),
                }
            )
            response = self.client.get(
                f'{reverse(self.export_email_view_name)}?{qs}',
                HTTP_REFERER=referer_url,
            )
            self.assertRedirects(response, referer_url)
            self.assertSpreadsheetEqual(
                mail.outbox[0].attachments[0][1],
                expected_spreadsheet_content,
                msg='Emailed contents do not match expected',
            )

    def test_export_no_data(self):
        """
        Test that the export view generates a spreadsheet without rows if the API call doesn't return any record.
        """
        expected_spreadsheet_content = [
            self.export_expected_xls_headers,
        ]

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)

            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )

            response = self.client.get(reverse(self.export_view_name))

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertSpreadsheetEqual(response.content, expected_spreadsheet_content)

    def test_email_export_no_data(self):
        """
        Test that the view sends the an empty spreadsheet via email.
        """
        expected_spreadsheet_content = [
            self.export_expected_xls_headers,
        ]

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)

            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )

            # get realistic referer
            qs = f'ordering={self.search_ordering}'
            response = self.client.get('/')
            referer_url = response.wsgi_request.build_absolute_uri(
                f'{reverse(self.view_name)}?{qs}',
            )

            response = self.client.get(
                f'{reverse(self.export_email_view_name)}?{qs}',
                HTTP_REFERER=referer_url,
            )
            self.assertRedirects(response, referer_url)
            self.assertSpreadsheetEqual(
                mail.outbox[0].attachments[0][1],
                expected_spreadsheet_content,
                msg='Emailed contents do not match expected',
            )

    def test_invalid_params_redirect_to_form(self):
        """
        Test that in case of invalid form, the export view redirects to the list page without
        taking any action.
        """
        # get realistic referer
        qs = 'ordering=invalid'
        response = self.client.get('/')
        referer_url = response.wsgi_request.build_absolute_uri(
            f'{reverse(self.view_name)}?{qs}',
        )

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            response = self.client.get(
                f'{reverse(self.export_email_view_name)}?{qs}',
                HTTP_REFERER=referer_url,
            )
            self.assertRedirects(response, referer_url)


class SenderViewsV2TestCase(
    SearchV2SecurityTestCaseMixin,
    ExportSecurityViewTestCaseMixin,
    SecurityViewTestCase,
):
    """
    Test case related to sender search V2 and detail views.
    """
    view_name = 'security:sender_list'
    advanced_search_view_name = 'security:sender_advanced_search'
    search_results_view_name = 'security:sender_search_results'
    detail_view_name = 'security:sender_detail'
    search_ordering = '-prisoner_count'
    api_list_path = '/senders/'

    export_view_name = 'security:sender_export'
    export_email_view_name = 'security:sender_email_export'
    export_expected_xls_headers = [
        'Sender name',
        'Payment method',
        'Credits sent',
        'Total amount sent',
        'Prisoners sent to',
        'Prisons sent to',
        'Bank transfer sort code',
        'Bank transfer account',
        'Bank transfer roll number',
        'Debit card number',
        'Debit card expiry',
        'Debit card postcode',
        'Other cardholder names',
        'Cardholder emails',
    ]
    export_expected_xls_rows = [
        [
            'MAISIE NOLAN',
            'Bank transfer',
            4,
            '£410.00',
            3,
            2,
            '10-10-10',
            '12312345',
            None,
            None,
            None,
            None,
            None,
            None,
        ],
        [
            'Maisie N',
            'Debit card',
            4,
            '£420.00',
            3,
            2,
            None,
            None,
            None,
            '************1234',
            '10/20',
            'SW137NJ',
            'Maisie Nolan',
            'm@outside.local, mn@outside.local',
        ]
    ]

    def get_api_object_list_response_data(self):
        return [
            self.bank_transfer_sender,
            self.debit_card_sender,
        ]

    def _test_search_results_content(self, response, advanced=False):
        response_content = response.content.decode(response.charset)

        self.assertIn('2 payment sources', response_content)
        self.assertIn('MAISIE NOLAN', response_content)
        self.assertIn('£410.00', response_content)
        self.assertIn('£420.00', response_content)

        if advanced:
            self.assertIn('Prison PRN', response_content)
            self.assertIn('Prison ABC', response_content)
        else:
            self.assertNotIn('Prison PRN', response_content)
            self.assertNotIn('Prison ABC', response_content)

    def test_detail_view_displays_bank_transfer_detail(self):
        sender_id = 9
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/senders/{sender_id}/'),
                json=self.bank_transfer_sender,
            )
            rsps.add(
                rsps.GET,
                api_url(f'/senders/{sender_id}/credits/'),
                json={
                    'count': 4,
                    'results': [self.credit_object] * 4,
                },
            )

            response = self.client.get(
                reverse(
                    self.detail_view_name,
                    kwargs={'sender_id': sender_id},
                ),
            )
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertIn('MAISIE', response_content)
        self.assertIn('12312345', response_content)
        self.assertIn('JAMES HALLS', response_content)
        self.assertIn('£102.50', response_content)

    def test_detail_view_displays_debit_card_detail(self):
        sender_id = 9
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/senders/{sender_id}/'),
                json=self.debit_card_sender,
            )
            rsps.add(
                rsps.GET,
                api_url(f'/senders/{sender_id}/credits/'),
                json={
                    'count': 4,
                    'results': [self.credit_object] * 4,
                }
            )
            response = self.client.get(
                reverse(
                    self.detail_view_name,
                    kwargs={'sender_id': sender_id},
                ),
            )
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertIn('************1234', response_content)
        self.assertIn('10/20', response_content)
        self.assertIn('SW13 7NJ', response_content)
        self.assertIn('JAMES HALLS', response_content)
        self.assertIn('£102.50', response_content)

    def test_detail_not_found(self):
        sender_id = 9
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/senders/{sender_id}/'),
                status=404,
            )
            with silence_logger('django.request'):
                response = self.client.get(
                    reverse(
                        self.detail_view_name,
                        kwargs={'sender_id': sender_id},
                    ),
                )
        self.assertEqual(response.status_code, 404)

    def test_connection_errors(self):
        sender_id = 9
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                status=500,
            )
            with silence_logger('django.request'):
                response = self.client.get(reverse(self.view_name))
        self.assertContains(response, 'non-field-error')

        with responses.RequestsMock() as rsps:
            no_saved_searches(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/senders/{sender_id}/'),
                status=500,
            )
            rsps.add(
                rsps.GET,
                api_url(f'/senders/{sender_id}/credits/'),
                status=500,
            )
            with silence_logger('django.request'):
                response = self.client.get(
                    reverse(
                        self.detail_view_name,
                        kwargs={'sender_id': sender_id},
                    ),
                )
        self.assertContains(response, 'non-field-error')


class PrisonerViewsV2TestCase(
    SearchV2SecurityTestCaseMixin,
    ExportSecurityViewTestCaseMixin,
    SecurityViewTestCase,
):
    """
    Test case related to prisoner search V2 and detail views.
    """
    view_name = 'security:prisoner_list'
    advanced_search_view_name = 'security:prisoner_advanced_search'
    search_results_view_name = 'security:prisoner_search_results'
    detail_view_name = 'security:prisoner_detail'
    search_ordering = '-sender_count'
    api_list_path = '/prisoners/'
    prison_api_filter_name = 'current_prison'

    export_view_name = 'security:prisoner_export'
    export_email_view_name = 'security:prisoner_email_export'
    export_expected_xls_headers = [
        'Prisoner number',
        'Prisoner name',
        'Date of birth',
        'Credits received',
        'Total amount received',
        'Payment sources',
        'Disbursements sent',
        'Total amount sent',
        'Recipients',
        'Current prison',
        'All known prisons',
        'Names given by senders',
    ]
    export_expected_xls_rows = [
        [
            'A1409AE',
            'JAMES HALLS',
            '1986-12-09',
            3,
            '£310.00',
            2,
            2,
            '£290.00',
            1,
            'Prison PRN',
            'Prison PRN',
            'Jim Halls, JAMES HALLS',
        ],
    ]

    def get_api_object_list_response_data(self):
        return [self.prisoner_profile]

    def _test_search_results_content(self, response, advanced=False):
        response_content = response.content.decode(response.charset)
        self.assertIn('JAMES HALLS', response_content)
        self.assertIn('A1409AE', response_content)
        self.assertIn('310.00', response_content)

        if advanced:
            self.assertIn('Prison PRN', response_content)
        else:
            self.assertNotIn('Prison PRN', response_content)

    def test_detail_view(self):
        prisoner_id = 9
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/prisoners/{prisoner_id}/'),
                json=self.prisoner_profile,
            )
            rsps.add(
                rsps.GET,
                api_url(f'/prisoners/{prisoner_id}/credits/'),
                json={
                    'count': 4,
                    'results': [self.credit_object] * 4,
                },
            )

            response = self.client.get(
                reverse(
                    self.detail_view_name,
                    kwargs={'prisoner_id': prisoner_id},
                ),
            )
        response_content = response.content.decode(response.charset)
        self.assertIn('JAMES HALLS', response_content)
        self.assertIn('Jim Halls', response_content)
        self.assertNotIn('James Halls', response_content)
        self.assertIn('MAISIE', response_content)
        self.assertIn('£102.50', response_content)

    def test_detail_not_found(self):
        prisoner_id = 999
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/prisoners/{prisoner_id}/'),
                status=404,
            )
            rsps.add(
                rsps.GET,
                api_url(f'/prisoners/{prisoner_id}/credits/'),
                status=404,
            )
            with silence_logger('django.request'):
                response = self.client.get(
                    reverse(
                        self.detail_view_name,
                        kwargs={'prisoner_id': prisoner_id},
                    ),
                )
        self.assertEqual(response.status_code, 404)

    def test_connection_errors(self):
        prisoner_id = 9
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                status=500,
            )
            with silence_logger('django.request'):
                response = self.client.get(reverse(self.view_name))
        self.assertContains(response, 'non-field-error')

        with responses.RequestsMock(assert_all_requests_are_fired=False) as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/prisoners/{prisoner_id}/'),
                status=500,
            )
            rsps.add(
                rsps.GET,
                api_url(f'/prisoners/{prisoner_id}/credits/'),
                status=500,
            )
            with silence_logger('django.request'):
                response = self.client.get(
                    reverse(
                        self.detail_view_name,
                        kwargs={'prisoner_id': prisoner_id},
                    ),
                )
        self.assertContains(response, 'non-field-error')


class CreditViewsV2TestCase(SearchV2SecurityTestCaseMixin, ExportSecurityViewTestCaseMixin, SecurityViewTestCase):
    """
    Test case related to credit search V2 and detail views.
    """
    view_name = 'security:credit_list'
    advanced_search_view_name = 'security:credit_advanced_search'
    search_results_view_name = 'security:credit_search_results'
    detail_view_name = 'security:credit_detail'
    search_ordering = '-received_at'
    api_list_path = '/credits/'

    debit_card_credit = {
        'id': 1,
        'source': 'online',
        'amount': 23000,
        'intended_recipient': 'Mr G Melley',
        'prisoner_number': 'A1411AE',
        'prisoner_name': 'GEORGE MELLEY',
        'prison': 'LEI',
        'prison_name': 'HMP LEEDS',
        'sender_name': None,
        'sender_sort_code': None,
        'sender_account_number': None,
        'sender_email': 'ian@mail.local',
        'billing_address': {'line1': '102PF', 'city': 'London'},
        'sender_roll_number': None,
        'card_number_first_digits': '111122',
        'card_number_last_digits': '4444',
        'card_expiry_date': '07/18',
        'resolution': 'credited',
        'owner': None,
        'owner_name': 'Maria',
        'started_at': '2016-05-25T20:21:00Z',
        'received_at': '2016-05-25T20:24:00Z',
        'credited_at': '2016-05-26T08:27:00Z',
        'refunded_at': None,
        'comments': [{'user_full_name': 'Eve', 'comment': 'OK'}],
        'nomis_transaction_id': None,
        'ip_address': '127.0.0.1',
    }
    bank_transfer_credit = {
        'id': 2,
        'source': 'bank_transfer',
        'amount': 27500,
        'intended_recipient': None,
        'prisoner_number': 'A1413AE',
        'prisoner_name': 'NORMAN STANLEY FLETCHER',
        'prison': 'LEI',
        'prison_name': 'HMP LEEDS',
        'sender_name': 'HEIDENREICH X',
        'sender_email': None,
        'billing_address': None,
        'sender_sort_code': '219657',
        'sender_account_number': '88447894',
        'card_number_last_digits': None,
        'card_expiry_date': None,
        'sender_roll_number': '',
        'resolution': 'credited',
        'owner': None,
        'owner_name': 'Maria',
        'started_at': None,
        'received_at': '2016-05-22T12:00:00Z',
        'credited_at': '2016-05-23T08:10:00Z',
        'refunded_at': None,
        'comments': [],
        'nomis_transaction_id': '123456-7',
        'ip_address': None,
    }

    export_view_name = 'security:credit_export'
    export_email_view_name = 'security:credit_email_export'
    export_expected_xls_headers = [
        'Internal ID',
        'Date started', 'Date received', 'Date credited',
        'Amount',
        'Prisoner number', 'Prisoner name', 'Prison',
        'Sender name', 'Payment method',
        'Bank transfer sort code', 'Bank transfer account', 'Bank transfer roll number',
        'Debit card number', 'Debit card expiry', 'Debit card billing address',
        'Sender email', 'Sender IP address',
        'Status',
        'NOMIS transaction',
    ]
    export_expected_xls_rows = [
        [
            1,
            '2016-05-25 21:21:00',  # in Europe/London
            '2016-05-25 21:24:00',
            '2016-05-26 09:27:00',
            '£230.00',
            'A1411AE',
            'GEORGE MELLEY',
            'HMP LEEDS',
            None,
            'Debit card',
            None,
            None,
            None,
            '111122******4444',
            '07/18',
            '102PF, London',
            'ian@mail.local',
            '127.0.0.1',
            'Credited',
            None,
        ],
        [
            2,
            None,
            '2016-05-22',
            '2016-05-23 09:10:00',
            '£275.00',
            'A1413AE',
            'NORMAN STANLEY FLETCHER',
            'HMP LEEDS',
            'HEIDENREICH X',
            'Bank transfer',
            '21-96-57',
            '88447894',
            None,
            None,
            None,
            None,
            None,
            None,
            'Credited',
            '123456-7',
        ],
    ]

    def get_api_object_list_response_data(self):
        return [
            self.debit_card_credit,
            self.bank_transfer_credit,
        ]

    def _test_search_results_content(self, response, advanced=False):
        response_content = response.content.decode(response.charset)
        self.assertIn('2 credits', response_content)

        self.assertIn('GEORGE MELLEY', response_content)
        self.assertIn('A1411AE', response_content)
        self.assertIn('230.00', response_content)

        self.assertIn('NORMAN STANLEY FLETCHER', response_content)
        self.assertIn('A1413AE', response_content)
        self.assertIn('275.00', response_content)

        # results page via advanced search includes an extra `prison` column
        if advanced:
            self.assertIn('HMP LEEDS', response_content)
        else:
            self.assertNotIn('HMP LEEDS', response_content)

    def test_detail_view_displays_debit_card_detail(self):
        credit_id = 2
        with responses.RequestsMock() as rsps:
            rsps.add(
                rsps.GET,
                f'{api_url(self.api_list_path)}?pk={credit_id}',
                json={
                    'count': 1,
                    'previous': None,
                    'next': None,
                    'results': [self.debit_card_credit],
                }
            )

            self.login(rsps)
            response = self.client.get(
                reverse(
                    self.detail_view_name,
                    kwargs={'credit_id': credit_id},
                ),
            )
        self.assertContains(response, 'Debit card')
        response_content = response.content.decode(response.charset)
        self.assertIn('£230.00', response_content)
        self.assertIn('GEORGE MELLEY', response_content)
        self.assertIn('Mr G Melley', response_content)
        self.assertIn('A1411AE', response_content)
        self.assertIn('Maria credited to NOMIS', response_content)
        self.assertIn('Eve', response_content)
        self.assertIn('OK', response_content)

    def test_detail_view_displays_bank_transfer_detail(self):
        credit_id = 2
        with responses.RequestsMock() as rsps:
            rsps.add(
                rsps.GET,
                f'{api_url(self.api_list_path)}?pk={credit_id}',
                json={
                    'count': 1,
                    'previous': None,
                    'next': None,
                    'results': [self.bank_transfer_credit],
                }
            )

            self.login(rsps)
            response = self.client.get(
                reverse(
                    self.detail_view_name,
                    kwargs={'credit_id': credit_id},
                ),
            )
        self.assertContains(response, 'Bank transfer')
        response_content = response.content.decode(response.charset)
        self.assertIn('£275.00', response_content)
        self.assertIn('NORMAN STANLEY FLETCHER', response_content)
        self.assertIn('21-96-57', response_content)
        self.assertIn('88447894', response_content)
        self.assertIn('Maria credited to NOMIS', response_content)

    def test_detail_not_found(self):
        credit_id = 999
        with responses.RequestsMock() as rsps:
            rsps.add(
                rsps.GET,
                f'{api_url(self.api_list_path)}?pk={credit_id}',
                json={
                    'count': 0,
                    'previous': None,
                    'next': None,
                    'results': [],
                }
            )

            self.login(rsps)
            with silence_logger('django.request'):
                response = self.client.get(
                    reverse(
                        self.detail_view_name,
                        kwargs={'credit_id': credit_id},
                    ),
                )
        self.assertEqual(response.status_code, 404)


class DisbursementViewsV2TestCase(
    SearchV2SecurityTestCaseMixin,
    ExportSecurityViewTestCaseMixin,
    SecurityViewTestCase,
):
    """
    Test case related to disbursement search V2 and detail views.
    """
    view_name = 'security:disbursement_list'
    advanced_search_view_name = 'security:disbursement_advanced_search'
    search_results_view_name = 'security:disbursement_search_results'
    detail_view_name = 'security:disbursement_detail'
    search_ordering = '-created'
    api_list_path = '/disbursements/'

    bank_transfer_disbursement = {
        'id': 99,
        'created': '2018-02-12T12:00:00Z',
        'modified': '2018-02-12T12:00:00Z',
        'method': 'bank_transfer',
        'amount': 2000,
        'resolution': 'sent',
        'nomis_transaction_id': '1234567-1',
        'invoice_number': '1000099',
        'prisoner_number': 'A1409AE',
        'prisoner_name': 'JAMES HALLS',
        'prison': 'ABC',
        'prison_name': 'HMP Test1',
        'recipient_first_name': 'Jack',
        'recipient_last_name': 'Halls',
        'recipient_email': '',
        'remittance_description': '',
        'address_line1': '102 Petty France',
        'address_line2': '',
        'city': 'London',
        'postcode': 'SW1H 9AJ',
        'country': None,
        'account_number': '1234567',
        'sort_code': '112233',
        'roll_number': None,
        'comments': [],
        'log_set': [
            {
                'action': 'created',
                'created': '2018-02-12T10:00:00Z',
                'user': {'first_name': 'Mary', 'last_name': 'Smith', 'username': 'msmith'}
            },
            {
                'action': 'confirmed',
                'created': '2018-02-12T11:00:00Z',
                'user': {'first_name': 'John', 'last_name': 'Smith', 'username': 'jsmith'}
            },
            {
                'action': 'sent',
                'created': '2018-02-12T12:00:00Z',
                'user': {'first_name': 'SSCL', 'last_name': '', 'username': 'sscl'}
            }
        ],
    }

    cheque_disbursement = {
        'id': 100,
        'created': '2018-02-10T10:00:00Z',
        'modified': '2018-02-10T12:00:00Z',
        'method': 'cheque',
        'amount': 1000,
        'resolution': 'confirmed',
        'nomis_transaction_id': '1234568-1',
        'invoice_number': None,
        'prisoner_number': 'A1401AE',
        'prisoner_name': 'JILLY HALL',
        'prison': 'DEF',
        'prison_name': 'HMP Test2',
        'recipient_first_name': 'Jilly',
        'recipient_last_name': 'Halls',
        'recipient_email': 'jilly@mtp.local',
        'remittance_description': 'PRESENT',
        'address_line1': '102 Petty France',
        'address_line2': '',
        'city': 'London',
        'postcode': 'SW1H 9AJ',
        'country': None,
        'account_number': '',
        'sort_code': '',
        'roll_number': None,
        'comments': [],
        'log_set': [
            {
                'action': 'created',
                'created': '2018-02-10T01:00:00Z',
                'user': {'first_name': 'Mary', 'last_name': 'Smith', 'username': 'msmith'}
            },
            {
                'action': 'confirmed',
                'created': '2018-02-10T02:00:00Z',
                'user': {'first_name': 'John', 'last_name': 'Smith', 'username': 'jsmith'}
            }
        ],
    }

    export_view_name = 'security:disbursement_export'
    export_email_view_name = 'security:disbursement_email_export'
    export_expected_xls_headers = [
        'Internal ID',
        'Date entered', 'Date confirmed', 'Date sent',
        'Amount',
        'Prisoner number', 'Prisoner name', 'Prison',
        'Recipient name', 'Payment method',
        'Bank transfer sort code', 'Bank transfer account', 'Bank transfer roll number',
        'Recipient address', 'Recipient email',
        'Status',
        'NOMIS transaction', 'SOP invoice number',
    ]
    export_expected_xls_rows = [
        [
            99,
            '2018-02-12 12:00:00',
            '2018-02-12 11:00:00',
            '2018-02-12 12:00:00',
            '£20.00',
            'A1409AE',
            'JAMES HALLS',
            'HMP Test1',
            'Jack Halls',
            'Bank transfer',
            '11-22-33',
            '1234567',
            None,
            '102 Petty France, London, SW1H 9AJ',
            None,
            'Sent',
            '1234567-1',
            '1000099',
        ],
        [
            100,
            '2018-02-10 10:00:00',
            '2018-02-10 02:00:00',
            None,
            '£10.00',
            'A1401AE',
            'JILLY HALL',
            'HMP Test2',
            'Jilly Halls',
            'Cheque',
            None,
            None,
            None,
            '102 Petty France, London, SW1H 9AJ',
            'jilly@mtp.local',
            'Confirmed',
            '1234568-1',
            None,
        ],
    ]

    def _test_search_results_content(self, response, advanced=False):
        response_content = response.content.decode(response.charset)
        self.assertIn('2 disbursements', response_content)

        self.assertIn('Jack Halls', response_content)
        self.assertIn('20.00', response_content)
        self.assertIn('A1409AE', response_content)

        self.assertIn('Jilly Halls', response_content)
        self.assertIn('10.00', response_content)
        self.assertIn('A1401AE', response_content)

        # results page via advanced search includes an extra `prison` column
        if advanced:
            self.assertIn('HMP Test1', response_content)
            self.assertIn('HMP Test2', response_content)
        else:
            self.assertNotIn('HMP Test1', response_content)
            self.assertNotIn('HMP Test2', response_content)

    def test_detail_view_displays_bank_transfer_detail(self):
        disbursement_id = 99
        with responses.RequestsMock() as rsps:
            rsps.add(
                rsps.GET,
                api_url(f'/disbursements/{disbursement_id}/'),
                json=self.bank_transfer_disbursement,
            )

            self.login(rsps)
            response = self.client.get(
                reverse(
                    self.detail_view_name,
                    kwargs={'disbursement_id': disbursement_id},
                ),
            )
        self.assertContains(response, 'Bank transfer')
        self.assertContains(response, '20.00')
        self.assertContains(response, 'JAMES HALLS')
        self.assertContains(response, 'Jack Halls')
        self.assertContains(response, '1234567-1')
        self.assertContains(response, '1000099')
        self.assertContains(response, 'Confirmed by John Smith')
        self.assertContains(response, 'None given')

    def test_detail_view_displays_cheque_detail(self):
        disbursement_id = 100
        with responses.RequestsMock() as rsps:
            rsps.add(
                rsps.GET,
                api_url(f'/disbursements/{disbursement_id}/'),
                json=self.cheque_disbursement,
            )

            self.login(rsps)
            response = self.client.get(
                reverse(
                    self.detail_view_name,
                    kwargs={'disbursement_id': disbursement_id},
                ),
            )
        self.assertContains(response, 'Cheque')
        self.assertContains(response, '10.00')
        self.assertContains(response, 'JILLY HALL')
        self.assertContains(response, 'Jilly Halls')
        self.assertContains(response, '1234568-1')
        self.assertContains(response, 'Confirmed by John Smith')
        self.assertContains(response, 'jilly@mtp.local')
        self.assertContains(response, 'PRESENT')

    def get_api_object_list_response_data(self):
        return [
            self.bank_transfer_disbursement,
            self.cheque_disbursement,
        ]


@contextmanager
def temp_spreadsheet(data):
    with tempfile.TemporaryFile() as f:
        f.write(data)
        wb = load_workbook(f)
        ws = wb.active
        yield ws


class PinnedProfileTestCase(SecurityViewTestCase):

    @responses.activate
    def test_pinned_profiles_on_dashboard(self):
        responses.add(
            responses.GET,
            api_url('/searches/'),
            json={
                'count': 2,
                'results': [
                    {
                        'id': 1,
                        'description': 'Saved search 1',
                        'endpoint': '/prisoners/1/credits',
                        'last_result_count': 2,
                        'site_url': '/en-gb/security/prisoners/1/',
                        'filters': []
                    },
                    {
                        'id': 2,
                        'description': 'Saved search 2',
                        'endpoint': '/senders/1/credits',
                        'last_result_count': 3,
                        'site_url': '/en-gb/security/senders/1/',
                        'filters': []
                    }
                ]
            },
        )
        responses.add(
            responses.GET,
            api_url('/prisoners/1/credits/'),
            json={
                'count': 5,
                'results': []
            },
        )
        responses.add(
            responses.GET,
            api_url('/senders/1/credits/'),
            json={
                'count': 10,
                'results': []
            },
        )
        response = self.login_test_searches(rsps=responses)

        self.assertContains(response, 'Saved search 1')
        self.assertContains(response, 'Saved search 2')
        self.assertContains(response, '3 new credits')
        self.assertContains(response, '7 new credits')

    @responses.activate
    def test_removes_invalid_saved_searches(self):
        responses.add(
            responses.GET,
            api_url('/searches/'),
            json={
                'count': 2,
                'results': [
                    {
                        'id': 1,
                        'description': 'Saved search 1',
                        'endpoint': '/prisoners/1/credits',
                        'last_result_count': 2,
                        'site_url': '/en-gb/security/prisoners/1/',
                        'filters': []
                    },
                    {
                        'id': 2,
                        'description': 'Saved search 2',
                        'endpoint': '/senders/1/credits',
                        'last_result_count': 3,
                        'site_url': '/en-gb/security/senders/1/',
                        'filters': []
                    }
                ]
            },
        )
        responses.add(
            responses.GET,
            api_url('/prisoners/1/credits/'),
            json={
                'count': 5,
                'results': []
            },
        )
        responses.add(
            responses.GET,
            api_url('/senders/1/credits/'),
            status=404,
        )
        responses.add(
            responses.DELETE,
            api_url('/searches/2/'),
            status=201,
        )
        response = self.login_test_searches(rsps=responses)

        self.assertContains(response, 'Saved search 1')
        self.assertNotContains(response, 'Saved search 2')
        self.assertContains(response, '3 new credits')


class PrisonerDetailViewTestCase(SecurityViewTestCase):
    def _add_prisoner_data_responses(self):
        responses.add(
            responses.GET,
            api_url('/prisoners/1/'),
            json=self.prisoner_profile,
        )
        responses.add(
            responses.GET,
            api_url('/prisoners/1/credits/'),
            json={
                'count': 4,
                'results': [
                    self.credit_object, self.credit_object,
                    self.credit_object, self.credit_object
                ],
            },
        )

    @responses.activate
    @mock.patch('security.views.nomis.can_access_nomis', mock.Mock(return_value=True))
    @mock.patch('security.views.nomis.get_photograph_data', mock.Mock(return_value=TEST_IMAGE_DATA))
    def test_display_nomis_photo(self):
        self.login(responses, follow=False)
        response = self.client.get(
            reverse(
                'security:prisoner_image',
                kwargs={'prisoner_number': self.prisoner_profile['prisoner_number']}
            )
        )
        self.assertContains(response, base64.b64decode(TEST_IMAGE_DATA))

    @responses.activate
    @mock.patch('security.views.nomis.can_access_nomis', mock.Mock(return_value=True))
    @mock.patch('security.views.nomis.get_photograph_data', mock.Mock(return_value=None))
    def test_missing_nomis_photo(self):
        self.login(responses, follow=False)
        response = self.client.get(
            reverse(
                'security:prisoner_image',
                kwargs={'prisoner_number': self.prisoner_profile['prisoner_number']}
            ),
            follow=False
        )
        self.assertRedirects(response, '/static/images/placeholder-image.png', fetch_redirect_response=False)

    @responses.activate
    def test_display_pinned_profile(self):
        self._add_prisoner_data_responses()
        responses.add(
            responses.GET,
            api_url('/searches/'),
            json={
                'count': 1,
                'results': [
                    {
                        'id': 1,
                        'description': 'Saved search 1',
                        'endpoint': '/prisoners/1/credits/',
                        'last_result_count': 2,
                        'site_url': '/en-gb/security/prisoners/1/?ordering=-received_at',
                        'filters': []
                    },
                ]
            },
        )
        responses.add(
            responses.PATCH,
            api_url('/searches/1/'),
            status=204,
        )
        self.login(responses, follow=False)
        response = self.client.get(
            reverse('security:prisoner_detail', kwargs={'prisoner_id': 1})
        )
        self.assertContains(response, 'Stop monitoring this prisoner')
        for call in responses.calls:
            if call.request.path_url == '/searches/1/':
                self.assertEqual(call.request.body, b'{"last_result_count": 4}')

    @responses.activate
    def test_pin_profile(self):
        self._add_prisoner_data_responses()
        responses.add(
            responses.GET,
            api_url('/searches/'),
            json={
                'count': 0,
                'results': []
            },
        )
        responses.add(
            responses.POST,
            api_url('/searches/'),
            status=201,
        )
        responses.add(
            responses.POST,
            api_url('/prisoners/1/monitor'),
            status=204,
        )

        self.login(responses, follow=False)
        self.client.get(
            reverse('security:prisoner_detail', kwargs={'prisoner_id': 1}) +
            '?pin=1'
        )

        for call in responses.calls:
            if call.request.path_url == '/searches/':
                self.assertEqual(
                    json.loads(call.request.body.decode()),
                    {
                        'description': 'A1409AE JAMES HALLS',
                        'endpoint': '/prisoners/1/credits/',
                        'last_result_count': 4,
                        'site_url': '/en-gb/security/prisoners/1/',
                        'filters': [{'field': 'ordering', 'value': '-received_at'}],
                    },
                )


class NotificationsTestCase(SecurityBaseTestCase):
    def login(self, rsps):
        super().login(
            rsps,
            user_data=self.get_user_data(
                flags=[
                    hmpps_employee_flag, confirmed_prisons_flag, provided_job_info_flag,
                ]
            )
        )

    def test_no_notifications_not_monitoring(self):
        """
        Expect to see a message if you're not monitoring anything and have no notifications
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/events/pages/') + '?rule=MONP&rule=MONS&offset=0&limit=25',
                json={'count': 0, 'newest': None, 'oldest': None},
                match_querystring=True
            )
            rsps.add(
                rsps.GET,
                api_url('/monitored/'),
                json={'count': 0},
            )
            rsps.add(
                rsps.GET,
                api_url('/events/'),
                json={'count': 0, 'results': []},
            )
            response = self.client.get(reverse('security:notification_list'))
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertIn('You’re not monitoring anything at the moment', response_content)
        self.assertIn('0 results', response_content)

    def test_no_notifications_but_monitoring(self):
        """
        Expect to see nothing interesting if monitoring some profile, but have no notifications
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/events/pages/') + '?rule=MONP&rule=MONS&offset=0&limit=25',
                json={'count': 0, 'newest': None, 'oldest': None},
                match_querystring=True
            )
            rsps.add(
                rsps.GET,
                api_url('/monitored/'),
                json={'count': 3},
            )
            rsps.add(
                rsps.GET,
                api_url('/events/'),
                json={'count': 0, 'results': []},
            )
            response = self.client.get(reverse('security:notification_list'))
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertNotIn('You’re not monitoring anything at the moment', response_content)
        self.assertIn('0 results', response_content)

    def test_notifications_not_monitoring(self):
        """
        Expect to see a message if you're not monitoring anything even if you have notifications
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/events/pages/') + '?rule=MONP&rule=MONS&offset=0&limit=25',
                json={'count': 1, 'newest': '2019-07-15', 'oldest': '2019-07-15'},
                match_querystring=True
            )
            rsps.add(
                rsps.GET,
                api_url('/monitored/'),
                json={'count': 0},
            )
            rsps.add(
                rsps.GET,
                api_url('/events/'),
                json={'count': 1, 'results': [
                    {
                        'id': 1,
                        'rule': 'MONP',
                        'triggered_at': '2019-07-15T10:00:00Z',
                        'credit_id': 1, 'disbursement_id': None,
                        'sender_profile': None, 'recipient_profile': None,
                        'prisoner_profile': {
                            'id': 1, 'prisoner_name': 'JAMES HALLS', 'prisoner_number': 'A1409AE',
                        },
                    }
                ]},
            )
            response = self.client.get(reverse('security:notification_list'))
            request_url = rsps.calls[-2].request.url
            request_query = request_url.split('?', 1)[1]
            request_query = QueryDict(request_query)
            self.assertEqual(request_query['triggered_at__gte'], '2019-07-15')
            self.assertEqual(request_query['triggered_at__lt'], '2019-07-16')
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertIn('You’re not monitoring anything at the moment', response_content)
        self.assertIn('1 result', response_content)
        self.assertIn('1 transaction', response_content)
        self.assertIn('JAMES HALLS (A1409AE)', response_content)

    def test_notifications_but_monitoring(self):
        """
        Expect to see a list of notifications when some exist
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/events/pages/') + '?rule=MONP&rule=MONS&offset=0&limit=25',
                json={'count': 1, 'newest': '2019-07-15', 'oldest': '2019-07-15'},
                match_querystring=True
            )
            rsps.add(
                rsps.GET,
                api_url('/monitored/'),
                json={'count': 3},
            )
            rsps.add(
                rsps.GET,
                api_url('/events/'),
                json={'count': 1, 'results': [
                    {
                        'id': 1,
                        'rule': 'MONP',
                        'triggered_at': '2019-07-15T10:00:00Z',
                        'credit_id': 1, 'disbursement_id': None,
                        'sender_profile': None, 'recipient_profile': None,
                        'prisoner_profile': {
                            'id': 1, 'prisoner_name': 'JAMES HALLS', 'prisoner_number': 'A1409AE',
                        },
                    }
                ]},
            )
            response = self.client.get(reverse('security:notification_list'))
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertNotIn('You’re not monitoring anything at the moment', response_content)
        self.assertIn('1 result', response_content)
        self.assertIn('1 transaction', response_content)
        self.assertIn('JAMES HALLS (A1409AE)', response_content)

    def test_notification_pages(self):
        """
        Expect the correct number of pages if there are many notifications
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/events/pages/') + '?rule=MONP&rule=MONS&offset=0&limit=25',
                json={'count': 32, 'newest': '2019-07-15', 'oldest': '2019-06-21'},
                match_querystring=True
            )
            rsps.add(
                rsps.GET,
                api_url('/monitored/'),
                json={'count': 3},
            )
            rsps.add(
                rsps.GET,
                api_url('/events/'),
                json={'count': 25, 'results': [
                    {
                        'id': 1 + days,
                        'rule': 'MONP',
                        'triggered_at': (
                            make_aware(datetime.datetime(2019, 7, 15, 10) - datetime.timedelta(days)).isoformat()
                        ),
                        'credit_id': 1 + days, 'disbursement_id': None,
                        'sender_profile': None, 'recipient_profile': None,
                        'prisoner_profile': {
                            'id': 1, 'prisoner_name': 'JAMES HALLS', 'prisoner_number': 'A1409AE',
                        },
                    }
                    for days in range(0, 25)
                ]},
            )
            response = self.client.get(reverse('security:notification_list'))
            request_url = rsps.calls[-2].request.url
            request_query = request_url.split('?', 1)[1]
            request_query = QueryDict(request_query)
            self.assertEqual(request_query['triggered_at__gte'], '2019-06-21')
            self.assertEqual(request_query['triggered_at__lt'], '2019-07-16')
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertIn('32 results', response_content)
        self.assertIn('Page 1 of 2.', response_content)

    def test_notification_grouping(self):
        """
        Expect notifications to be grouped by connected profile
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/events/pages/') + '?rule=MONP&rule=MONS&offset=0&limit=25',
                json={'count': 1, 'newest': '2019-07-15', 'oldest': '2019-07-15'},
                match_querystring=True
            )
            rsps.add(
                rsps.GET,
                api_url('/monitored/'),
                json={'count': 4},
            )
            rsps.add(
                rsps.GET,
                api_url('/events/'),
                json={'count': 6, 'results': [
                    {
                        'id': 1,
                        'rule': 'MONP',
                        'triggered_at': '2019-07-15T10:00:00Z',
                        'credit_id': 1, 'disbursement_id': None,
                        'sender_profile': None, 'recipient_profile': None,
                        'prisoner_profile': {
                            'id': 1, 'prisoner_name': 'JAMES HALLS', 'prisoner_number': 'A1409AE',
                        },
                    },
                    {
                        'id': 2,
                        'rule': 'MONS',
                        'triggered_at': '2019-07-15T09:00:00Z',
                        'credit_id': 2, 'disbursement_id': None,
                        'prisoner_profile': None, 'recipient_profile': None,
                        'sender_profile': {
                            'id': 1, 'bank_transfer_details': [{'sender_name': 'Mary Halls'}],
                        },
                    },
                    {
                        'id': 3,
                        'rule': 'MONP',
                        'triggered_at': '2019-07-15T08:00:00Z',
                        'credit_id': 3, 'disbursement_id': None,
                        'sender_profile': None, 'recipient_profile': None,
                        'prisoner_profile': {
                            'id': 2, 'prisoner_name': 'JILLY HALL', 'prisoner_number': 'A1401AE',
                        },
                    },
                    {
                        'id': 4,
                        'rule': 'MONP',
                        'triggered_at': '2019-07-15T07:00:00Z',
                        'credit_id': None, 'disbursement_id': 1,
                        'sender_profile': None, 'recipient_profile': None,
                        'prisoner_profile': {
                            'id': 1, 'prisoner_name': 'JAMES HALLS', 'prisoner_number': 'A1409AE',
                        },
                    },
                    {
                        'id': 5,
                        'rule': 'MONS',
                        'triggered_at': '2019-07-15T06:00:00Z',
                        'credit_id': 5, 'disbursement_id': None,
                        'prisoner_profile': None, 'recipient_profile': None,
                        'sender_profile': {
                            'id': 1, 'bank_transfer_details': [{'sender_name': 'Mary Halls'}],
                        },
                    },
                    {
                        'id': 6,
                        'rule': 'MONS',
                        'triggered_at': '2019-07-15T05:00:00Z',
                        'credit_id': 6, 'disbursement_id': None,
                        'prisoner_profile': None, 'recipient_profile': None,
                        'sender_profile': {
                            'id': 2, 'debit_card_details': [{'cardholder_names': ['Fred Smith']}],
                        },
                    },
                ]},
            )
            response = self.client.get(reverse('security:notification_list'))
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertIn('Page 1 of 1.', response_content)
        self.assertEqual(response_content.count('Mary Halls'), 1)
        self.assertEqual(response_content.count('Fred Smith'), 1)
        self.assertEqual(response_content.count('JAMES HALLS'), 1)
        self.assertEqual(response_content.count('JILLY HALL'), 1)
        self.assertEqual(response_content.count('1 transaction'), 2)
        self.assertEqual(response_content.count('2 transactions'), 2)


class SettingsTestCase(SecurityBaseTestCase):
    def test_can_turn_on_email_notifications_switch(self):
        with responses.RequestsMock() as rsps:
            self.login(
                rsps,
                user_data=self.get_user_data(flags=[
                    hmpps_employee_flag, confirmed_prisons_flag,
                    provided_job_info_flag,
                ])
            )
            rsps.add(
                rsps.GET,
                api_url('/emailpreferences/'),
                json={'frequency': EmailNotifications.never},
            )
            response = self.client.get(reverse('settings'), follow=True)
            self.assertContains(response, 'not currently receiving email notifications')

            rsps.add(
                rsps.POST,
                api_url('/emailpreferences/'),
            )
            rsps.replace(
                rsps.GET,
                api_url('/emailpreferences/'),
                json={'frequency': EmailNotifications.daily},
            )
            response = self.client.post(reverse('settings'), data={'email_notifications': 'True'}, follow=True)
            self.assertNotContains(response, 'not currently receiving email notifications')

            last_post_call = list(filter(lambda call: call.request.method == rsps.POST, rsps.calls))[-1]
            last_request_body = json.loads(last_post_call.request.body)
            self.assertDictEqual(last_request_body, {'frequency': 'daily'})

    def test_can_turn_off_email_notifications_switch(self):
        with responses.RequestsMock() as rsps:
            self.login(
                rsps,
                user_data=self.get_user_data(
                    flags=[
                        hmpps_employee_flag, confirmed_prisons_flag,
                        provided_job_info_flag,
                    ]
                )
            )
            rsps.add(
                rsps.GET,
                api_url('/emailpreferences/'),
                json={'frequency': EmailNotifications.daily},
            )
            response = self.client.get(reverse('settings'), follow=True)
            self.assertNotContains(response, 'not currently receiving email notifications')

            rsps.add(
                rsps.POST,
                api_url('/emailpreferences/'),
            )
            rsps.replace(
                rsps.GET,
                api_url('/emailpreferences/'),
                json={'frequency': EmailNotifications.never},
            )
            response = self.client.post(reverse('settings'), data={'email_notifications': 'False'}, follow=True)
            self.assertContains(response, 'not currently receiving email notifications')

            last_post_call = list(filter(lambda call: call.request.method == rsps.POST, rsps.calls))[-1]
            last_request_body = json.loads(last_post_call.request.body)
            self.assertDictEqual(last_request_body, {'frequency': 'never'})


class LegacyViewsRedirectTestCase(SecurityBaseTestCase):
    """
    Tests related to legacy views.
    """

    @responses.activate
    def test_legacy_search_views_redirect_to_new_ones(self):
        """
        Test that legacy search views redirect to search views V2.
        """
        view_names = (
            'security:credit_list',
            'security:disbursement_list',
            'security:sender_list',
            'security:prisoner_list',
        )
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            mock_prison_response(rsps=rsps)

            for view_name in view_names:
                response = self.client.get(reverse(f'{view_name}_legacy'))
                self.assertRedirects(response, reverse(view_name))


class BaseCheckViewTestCase(SecurityBaseTestCase):
    """
    Tests related to the check views.
    """
    sender_id = 2
    prisoner_id = 3
    credit_id = 5

    credit_created_date = datetime.datetime.now()

    SAMPLE_CHECK_BASE = {
        'id': 1,
        'description': 'lorem ipsum',
        'rules': ['RULE1', 'RULE2'],
        'status': 'pending',
        'actioned_at': None,
        'actioned_by': None,
        'assigned_to': 7,
        'auto_accept_rule_state': None
    }

    SAMPLE_CREDIT_BASE = {
        'id': 1,
        'amount': 1000,
        'card_expiry_date': '02/20',
        'card_number_first_digits': '123456',
        'card_number_last_digits': '9876',
        'prisoner_name': 'Jean Valjean',
        'prisoner_number': '24601',
        'sender_email': 'sender@example.com',
        'sender_name': 'MAISIE NOLAN',
        'source': 'online',
        'started_at': '2019-07-02T10:00:00Z',
        'received_at': None,
        'credited_at': None,
        'prisoner_profile': prisoner_id,
        'billing_address': {
            'debit_card_sender_details': 17
        }
    }

    SAMPLE_CHECK = dict(SAMPLE_CHECK_BASE, credit=SAMPLE_CREDIT_BASE)

    SENDER_CREDIT = dict(
        SAMPLE_CREDIT_BASE,
        **{
            'security_check': SAMPLE_CHECK_BASE.copy(),
            'intended_recipient': 'Mr G Melley',
            'prisoner_name': 'Ms A. Nother Prisoner',
            'prisoner_number': 'Number 6',
            'amount': 1000000,
            'prison': 'LEI',
            'prison_name': 'HMP LEEDS',
            'billing_address': {'line1': '102PF', 'city': 'London'},
            'resolution': 'rejected',
            'started_at': credit_created_date.isoformat()
        }
    )
    SENDER_CREDIT['security_check']['description'] = ['Strict compliance check failed']
    SENDER_CREDIT['security_check']['actioned_by_name'] = 'Javert'
    SENDER_CREDIT['security_check']['actioned_at'] = credit_created_date.isoformat()
    SENDER_CREDIT['security_check']['status'] = 'rejected'

    SENDER_CHECK = copy.deepcopy(SAMPLE_CHECK)
    SENDER_CHECK['credit']['sender_profile'] = sender_id
    SENDER_CHECK['credit']['prisoner_profile'] = prisoner_id
    SENDER_CHECK['credit']['id'] = credit_id
    SENDER_CHECK['credit']['billing_address'] = {'debit_card_sender_details': 42}

    SENDER_CHECK_REJECTED = dict(SENDER_CHECK, status='rejected')

    PRISONER_CREDIT = dict(
        SAMPLE_CREDIT_BASE,
        **{
            'security_check': SAMPLE_CHECK_BASE.copy(),
            'amount': 10,
            'card_expiry_date': '02/50',
            'card_number_first_digits': '01199988199',
            'card_number_last_digits': '7253',
            'sender_email': 'someoneelse@example.com',
            'sender_name': 'SOMEONE ELSE',
            'prison': 'LEI',
            'prison_name': 'HMP LEEDS',
            'billing_address': {'line1': 'Somewhere else', 'city': 'London'},
            'resolution': 'credited',
        }
    )
    PRISONER_CREDIT['security_check']['description'] = ['Soft compliance check failed']
    PRISONER_CREDIT['security_check']['actioned_at'] = credit_created_date.isoformat()
    PRISONER_CREDIT['security_check']['actioned_by_name'] = 'Staff'
    PRISONER_CREDIT['security_check']['status'] = 'accepted'

    required_checks_permissions = (
        *required_permissions,
        'security.view_check',
        'security.change_check',
    )

    def get_user_data(
        self,
        *args,
        permissions=required_checks_permissions,
        **kwargs,
    ):
        """
        Adds extra permissions to manage checks.
        """
        return super().get_user_data(*args, permissions=permissions, **kwargs)

    def mock_need_attention_count(self, rsps, count):
        rsps.add(
            rsps.GET,
            api_url('/security/checks/'),
            json={
                'count': count,
                'results': [self.SAMPLE_CHECK] * count,
            }
        )

    @classmethod
    def _get_prisoner_credit_list(cls, length):
        for i in range(length):
            yield dict(cls.PRISONER_CREDIT, id=i)

    @classmethod
    def _get_sender_credit_list(cls, length):
        for i in range(length):
            yield dict(cls.SENDER_CREDIT, id=i)


class CheckListViewTestCase(BaseCheckViewTestCase):
    """
    Tests related to CheckListView.
    """

    def test_cannot_access_view(self):
        """
        Test that if the logged in user doesn't have the right permissions, he/she
        gets redirected to the dashboard.
        """
        user_data = self.get_user_data(permissions=required_permissions)
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps, user_data=user_data)
            response = self.client.get(reverse('security:check_list'), follow=True)
            self.assertRedirects(response, reverse('security:dashboard'))

    @mock.patch('security.forms.check.get_need_attention_date')
    def test_view(self, mock_get_need_attention_date):
        """
        Test that the view displays the pending checks returned by the API.
        """
        mock_get_need_attention_date.return_value = make_aware(datetime.datetime(2019, 7, 2, 9))

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            self.mock_need_attention_count(rsps, 0)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 1,
                    'results': [self.SAMPLE_CHECK],
                },
            )

            response = self.client.get(reverse('security:check_list'), follow=True)
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')

            content = response.content.decode()
            self.assertIn('24601', content)
            self.assertIn('1 credit', content)
            self.assertIn('This credit does not need attention today', content)
            self.assertNotIn('credit needs attention', content)
            self.assertNotIn('credits need attention', content)

    @mock.patch('security.forms.check.get_need_attention_date')
    def test_displays_count_of_credits_needing_attention(self, mock_get_need_attention_date):
        """
        Test that the view shows how many credits need attention.
        """
        mock_get_need_attention_date.return_value = make_aware(datetime.datetime(2019, 7, 9, 9))

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            self.mock_need_attention_count(rsps, 2)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 2,
                    'results': [self.SAMPLE_CHECK, self.SAMPLE_CHECK],
                },
            )

            response = self.client.get(reverse('security:check_list'), follow=True)
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')

            content = response.content.decode()
            self.assertIn('2 credits need attention', content)
            self.assertIn('This credit needs attention today!', content)

    def test_calculation_of_date_before_which_checks_need_attention(self):
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            self.mock_need_attention_count(rsps, 0)
            response = self.client.get(reverse('security:check_list'))
            form = response.context['form']
            api_call_made = rsps.calls[-3].request.url
            parsed_qs = parse_qs(api_call_made.split('?', 1)[1])
            self.assertEqual(parsed_qs['started_at__lt'], [form.need_attention_date.strftime('%Y-%m-%d %H:%M:%S')])

    def test_credit_row_has_review_link(self):
        """
        Test that the view displays link to review a credit.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            self.mock_need_attention_count(rsps, 0)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 1,
                    'results': [self.SAMPLE_CHECK],
                },
            )

            response = self.client.get(reverse('security:check_list'), follow=True)

            self.assertContains(response, 'Review <span class="visually-hidden">credit to Jean Valjean</span>')


class MyCheckListViewTestCase(BaseCheckViewTestCase):
    """
    Tests related to CheckListView.
    """

    def test_cannot_access_view(self):
        """
        Test that if the logged in user doesn't have the right permissions, he/she
        gets redirected to the dashboard.
        """
        user_data = self.get_user_data(permissions=required_permissions)
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps, user_data=user_data)
            response = self.client.get(reverse('security:my_check_list'), follow=True)
            self.assertRedirects(response, reverse('security:dashboard'))

    @mock.patch('security.forms.check.get_need_attention_date')
    def test_view(self, mock_get_need_attention_date):
        """
        Test that the view displays the pending checks returned by the API.
        """
        mock_get_need_attention_date.return_value = make_aware(datetime.datetime(2019, 7, 2, 9))

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            self.mock_need_attention_count(rsps, 0)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 1,
                    'results': [self.SAMPLE_CHECK],
                },
            )

            response = self.client.get(reverse('security:my_check_list'), follow=True)
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')

            content = response.content.decode()
            self.assertIn('24601', content)
            self.assertIn('1 credit', content)
            self.assertIn('This credit does not need attention today', content)
            self.assertNotIn('credit needs attention', content)
            self.assertNotIn('credits need attention', content)
            self.assertIn('My list (1)', content)

    @mock.patch('security.forms.check.get_need_attention_date')
    def test_displays_count_of_credits_needing_attention(self, mock_get_need_attention_date):
        """
        Test that the view shows how many credits need attention.
        """
        mock_get_need_attention_date.return_value = make_aware(datetime.datetime(2019, 7, 9, 9))

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            self.mock_need_attention_count(rsps, 2)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 2,
                    'results': [self.SAMPLE_CHECK, self.SAMPLE_CHECK],
                },
            )

            response = self.client.get(reverse('security:my_check_list'), follow=True)
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')

            content = response.content.decode()
            self.assertIn('2 credits need attention', content)
            self.assertIn('This credit needs attention today!', content)

    def test_calculation_of_date_before_which_checks_need_attention(self):
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            self.mock_need_attention_count(rsps, 0)
            response = self.client.get(reverse('security:my_check_list'))
            form = response.context['form']
            api_call_made = rsps.calls[-3].request.url
            parsed_qs = parse_qs(api_call_made.split('?', 1)[1])
            self.assertEqual(parsed_qs['started_at__lt'], [form.need_attention_date.strftime('%Y-%m-%d %H:%M:%S')])

    def test_credit_row_has_review_link(self):
        """
        Test that the view displays link to review a credit.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            self.mock_need_attention_count(rsps, 0)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 1,
                    'results': [self.SAMPLE_CHECK],
                },
            )

            response = self.client.get(reverse('security:my_check_list'), follow=True)

            self.assertContains(response, 'Review <span class="visually-hidden">credit to Jean Valjean</span>')


class CreditsHistoryListViewTestCase(BaseCheckViewTestCase):
    """
    Tests related to CreditsHistoryListView.
    """
    SAMPLE_CHECK_WITH_ACTIONED_BY = dict(
        BaseCheckViewTestCase.SAMPLE_CHECK,
        **{
            'actioned_at': '2020-01-13 12:00:00+00',
            'actioned_by': 1,
            'decision_reason': 'Money issues',
            'actioned_by_name': 'Barry Garlow',
            'status': 'accepted'
        }
    )

    SAMPLE_CHECK_WITH_ACTIONED_BY['credit'].update(
        {
            'billing_address': {
                'id': 21,
                'line1': 'Studio 33',
                'line2': 'Allen port',
                'city': 'Gloverside',
                'country': 'UK',
                'postcode': 'S1 3HS',
                'debit_card_sender_details': 17
            },
            'prison_name': 'Brixton Prison',
            'prisoner_number': '24601',
        }
    )

    def test_cannot_access_view(self):
        """
        Test that if the logged in user doesn't have the right permissions, he/she
        gets redirected to the dashboard.
        """
        user_data = self.get_user_data(permissions=required_permissions)
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps, user_data=user_data)
            response = self.client.get(reverse('security:credits_history'), follow=True)
            self.assertRedirects(response, reverse('security:dashboard'))

    def test_view(self):
        """
        Test that the view displays the history of checks caught by delayed capture by the API.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/security/checks/?{querystring}'.format(
                        querystring=urlencode([
                            ('started_at__gte', '2020-01-02T12:00:00'),
                            ('actioned_by', True),
                            ('offset', 0),
                            ('limit', 20),
                            ('ordering', '-created')
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 1,
                    'results': [self.SAMPLE_CHECK_WITH_ACTIONED_BY],
                },
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/security/checks/?{querystring}'.format(
                        querystring=urlencode([
                            ('status', 'pending'),
                            ('credit_resolution', 'initial'),
                            ('assigned_to', 5),
                            ('offset', 0),
                            ('limit', 1)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 2,
                    'results': [self.SAMPLE_CHECK_WITH_ACTIONED_BY, self.SAMPLE_CHECK_WITH_ACTIONED_BY],
                }
            )

            response = self.client.get(reverse('security:credits_history'))
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')

            content = response.content.decode()

            self.assertIn('24601', content)
            self.assertIn('1 credit', content)
            self.assertIn('My list (2)', content)

    def test_view_contains_relevant_data(self):
        """
        Test that the view displays the correct data from the API.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 1,
                    'results': [self.SAMPLE_CHECK_WITH_ACTIONED_BY],
                },
            )
            response = self.client.get(reverse('security:credits_history'))
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')
            self.assertContains(response, 'Barry Garlow')
            self.assertContains(response, 'Money issues')
            self.assertContains(response, 'S1 3HS')
            self.assertContains(response, '24601')
            self.assertContains(response, 'accepted')
            self.assertContains(response, 'Brixton Prison')
            self.assertContains(response, 'Decision details:')

    def test_view_does_not_display_decision_if_none(self):
        """
        Test that the view displays the correct data from the API.
        """
        self.SAMPLE_CHECK_WITH_ACTIONED_BY['decision_reason'] = None

        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 1,
                    'results': [self.SAMPLE_CHECK_WITH_ACTIONED_BY],
                },
            )
            response = self.client.get(reverse('security:credits_history'))

            self.assertContains(response, 'No decision reason entered')

    @parameterized.expand(
        CHECK_REJECTION_CATEGORY_BOOLEAN_MAPPING.items()
    )
    def test_credit_history_row_has_reason_checkbox_populated(
        self, rejection_reason_key, rejection_reason_full
    ):
        """
        Test that the view displays checkboxes associated with a credit.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/security/checks/?{querystring}'.format(
                        querystring=urlencode([
                            ('started_at__gte', '2020-01-02T12:00:00'),
                            ('actioned_by', True),
                            ('offset', 0),
                            ('limit', 20),
                            ('ordering', '-created')
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 1,
                    'results': [
                        dict(
                            self.SENDER_CHECK_REJECTED,
                            rejection_reasons={rejection_reason_key: True}
                        )
                    ]
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/security/checks/?{querystring}'.format(
                        querystring=urlencode([
                            ('status', 'pending'),
                            ('credit_resolution', 'initial'),
                            ('assigned_to', 5),
                            ('offset', 0),
                            ('limit', 1)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 2,
                    'results': [self.SAMPLE_CHECK_WITH_ACTIONED_BY, self.SAMPLE_CHECK_WITH_ACTIONED_BY],
                }
            )

            response = self.client.get(reverse('security:credits_history'))

            self.assertContains(response, rejection_reason_full)

    @parameterized.expand(
        generate_rejection_category_test_cases_text()
    )
    def test_credit_history_row_has_string_reason_populated(
        self, rejection_reason_key, rejection_reason_value
    ):
        """
        Test that the view displays the reason for associated populated checkboxes with a credit.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/security/checks/?{querystring}'.format(
                        querystring=urlencode([
                            ('started_at__gte', '2020-01-02T12:00:00'),
                            ('actioned_by', True),
                            ('offset', 0),
                            ('limit', 20),
                            ('ordering', '-created')
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 1,
                    'results': [
                        dict(
                            self.SENDER_CHECK_REJECTED,
                            rejection_reasons={rejection_reason_key: rejection_reason_value}
                        )
                    ]
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/security/checks/?{querystring}'.format(
                        querystring=urlencode([
                            ('status', 'pending'),
                            ('credit_resolution', 'initial'),
                            ('assigned_to', 5),
                            ('offset', 0),
                            ('limit', 1)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 2,
                    'results': [self.SAMPLE_CHECK_WITH_ACTIONED_BY, self.SAMPLE_CHECK_WITH_ACTIONED_BY],
                }
            )

            response = self.client.get(reverse('security:credits_history'))

            self.assertContains(response, rejection_reason_value)


class AcceptOrRejectCheckViewTestCase(BaseCheckViewTestCase, SecurityViewTestCase):
    """
    Tests related to AcceptOrRejectCheckView.
    """

    def test_cannot_access_view(self):
        """
        Test that if the logged in user doesn't have the right permissions, he/she
        gets redirected to the dashboard.
        """
        user_data = self.get_user_data(permissions=required_permissions)
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps, user_data=user_data)
            url = reverse('security:resolve_check', kwargs={'check_id': 1})
            response = self.client.get(url, follow=True)
            self.assertRedirects(response, reverse('security:dashboard'))

    def test_get(self):
        """
        Test that the view displays the pending check returned by the API.
        """
        response_len = 4
        check_id = 1
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_prisoner_credit_list(response_len))
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.get(url, follow=True)

            self.assertContains(response, 'Accept credit')
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')

    def test_get_with_previous_unbound_active_auto_accept(self):
        """
        Test that the view renders the form without auto-accept form
        """
        response_len = 4
        check_id = 1
        reason = 'Prisoners mother'
        different_user_data = self.get_user_data(first_name='different', last_name='user')
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_prisoner_credit_list(response_len))
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 1, 'prev': None, 'next': None, 'results': [{
                    'debit_card_sender_details': self.SENDER_CHECK['credit']['billing_address'][
                        'debit_card_sender_details'
                    ],
                    'prisoner_profile': self.SENDER_CHECK['credit']['prisoner_profile'],
                    'states': [
                        {
                            'created': (datetime.datetime.now(tz=pytz.utc) - datetime.timedelta(hours=3)).isoformat(),
                            'active': True
                        },
                        {
                            'created': (datetime.datetime.now(tz=pytz.utc) - datetime.timedelta(hours=2)).isoformat(),
                            'active': False
                        },
                        {
                            'created': (datetime.datetime.now(tz=pytz.utc) - datetime.timedelta(hours=1)).isoformat(),
                            'active': True,
                            'reason': reason,
                            'added_by': {
                                'first_name': different_user_data['first_name'],
                                'last_name': different_user_data['last_name'],
                            }
                        },
                    ]
                }]}
            )

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.get(url, follow=True)

            self.assertContains(response, 'Accept credit')
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')
            self.assertNotContains(response, 'Automatically accept future credits from')
            self.assertContains(
                response,
                f'Auto accept started for credits from {self.SENDER_CHECK["credit"]["sender_name"]} to '
                f'{self.SENDER_CHECK["credit"]["prisoner_number"]}'
            )
            self.assertContains(response, 'Started by:')
            self.assertContains(response, f'{different_user_data["first_name"]} {different_user_data["last_name"]}')
            self.assertContains(response, 'Date:')
            self.assertContains(
                response,
                (datetime.datetime.now(tz=pytz.utc) - datetime.timedelta(hours=1)).strftime('%d/%m/%Y %H:%M')
            )
            self.assertContains(response, 'Reason for automatically accepting:')
            self.assertContains(response, reason)

    def test_get_with_previous_unbound_inactive_auto_accept(self):
        """
        Test that the view renders the form with auto-accept form
        """
        response_len = 4
        check_id = 1
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_prisoner_credit_list(response_len))
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 1, 'prev': None, 'next': None, 'results': [{
                    'debit_card_sender_details': self.SENDER_CHECK['credit']['billing_address'][
                        'debit_card_sender_details'
                    ],
                    'prisoner_profile': self.SENDER_CHECK['credit']['prisoner_profile'],
                    'states': [
                        {
                            'created': (datetime.datetime.now() - datetime.timedelta(hours=2)).isoformat(),
                            'active': True
                        },
                        {
                            'created': (datetime.datetime.now() - datetime.timedelta(hours=1)).isoformat(),
                            'active': False
                        },
                    ]
                }]}
            )

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.get(url, follow=True)

            self.assertContains(response, 'Accept credit')
            self.assertContains(response, '123456******9876')
            self.assertContains(response, '02/20')
            self.assertContains(response, 'Automatically accept future credits from')
            self.assertNotContains(
                response,
                f'Auto accept started for credits from {self.SENDER_CHECK["credit"]["sender_name"]} to '
                f'{self.SENDER_CHECK["credit"]["prisoner_number"]}'
            )

    def test_check_view_hides_action_buttons_if_resolved_already(self):
        """
        There is currently nothing to prevent the view from showing a resolved security check.
        Test that the accept/reject form is absent for resolved checks.
        NB: as this test is not concerned with the api being called with the right parameters, the responses mock
            does not force the query string to be correct unlike other tests.
        """
        response_len = 0
        check_id = 1
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK_REJECTED
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/'.format(
                        sender_profile_id=self.sender_id,
                    ),
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/'.format(
                        prisoner_profile_id=self.prisoner_id,
                    ),
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_prisoner_credit_list(response_len))
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.get(url)
        content = response.content.decode()
        self.assertNotIn('name="fiu_action"', content, msg='Action button exists on page')
        self.assertIn('This credit was rejected by', content)

    def test_check_view_includes_matching_credit_history(self):
        """
        Test that the view displays credits related by sender id to the credit subject to a check.
        """
        check_id = 1
        response_len = 4
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len))
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True),
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_prisoner_credit_list(response_len))
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            response = self.client.get(
                reverse(
                    'security:resolve_check',
                    kwargs={'check_id': check_id},
                ),
            )
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertIn('123456******9876', response_content)
        self.assertIn('02/20', response_content)
        self.assertIn('Jean Valjean', response_content)
        self.assertIn('£10.00', response_content)

        # Senders previous credit
        self.assertIn('Ms A. Nother Prisoner', response_content)
        self.assertIn('£10,000.00', response_content)
        self.assertIn('Strict compliance check failed', response_content)
        self.assertIn('Javert', response_content)

        # Prisoners previous credit
        self.assertIn('£0.10', response_content)
        self.assertIn('01199988199******7253', response_content)
        self.assertIn('02/50', response_content)
        self.assertIn('SOMEONE ELSE', response_content)
        self.assertIn('Number 6', response_content)
        self.assertIn('Soft compliance check failed', response_content)
        self.assertIn('Staff', response_content)

        # TODO add in assertion for ordering

    def test_check_view_includes_matching_credit_history_including_active_auto_accept(self):
        """
        Test that the view displays auto-accepted credits related by sender id to the credit subject to a check.
        """
        check_id = 1
        response_len = 2
        sender_active_auto_accept_rule_state = {
            'active': True,
            'reason': 'I should be here sender'
        }
        sender_inactive_auto_accept_rule_state = {
            'active': False,
            'reason': 'I shouldnt be here sender'
        }
        prisoner_active_auto_accept_rule_state = {
            'active': True,
            'reason': 'I should be here prisoner'
        }
        prisoner_inactive_auto_accept_rule_state = {
            'active': False,
            'reason': 'I shouldnt be here prisoner'
        }
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 2,
                    'results': [
                        dict(
                            self.SENDER_CREDIT,
                            id=0,
                            security_check=dict(
                                self.PRISONER_CREDIT['security_check'],
                                auto_accept_rule_state=sender_active_auto_accept_rule_state
                            )
                        ),
                        dict(
                            self.SENDER_CREDIT,
                            id=1,
                            security_check=dict(
                                self.SENDER_CREDIT['security_check'],
                                auto_accept_rule_state=sender_inactive_auto_accept_rule_state
                            )
                        )
                    ]
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True),
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 2,
                    'results': [
                        dict(
                            self.PRISONER_CREDIT,
                            id=0,
                            security_check=dict(
                                self.PRISONER_CREDIT['security_check'],
                                auto_accept_rule_state=prisoner_active_auto_accept_rule_state
                            )
                        ),
                        dict(
                            self.PRISONER_CREDIT,
                            id=1,
                            security_check=dict(
                                self.PRISONER_CREDIT['security_check'],
                                auto_accept_rule_state=prisoner_inactive_auto_accept_rule_state
                            )
                        )
                    ]
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept?{}/'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            response = self.client.get(
                reverse(
                    'security:resolve_check',
                    kwargs={'check_id': check_id},
                ),
            )
        self.assertEqual(response.status_code, 200)
        response_content = response.content.decode(response.charset)
        self.assertIn('123456******9876', response_content)
        self.assertIn('02/20', response_content)
        self.assertIn('Jean Valjean', response_content)
        self.assertIn('£10.00', response_content)

        # Senders previous credit
        self.assertIn('Ms A. Nother Prisoner', response_content)
        self.assertIn('£10,000.00', response_content)
        self.assertIn('Strict compliance check failed', response_content)
        self.assertIn('Javert', response_content)
        self.assertIn(
            'Reason for automatically accepting:',
            response_content
        )
        self.assertIn(
            sender_active_auto_accept_rule_state['reason'],
            response_content
        )
        self.assertNotIn(
            sender_inactive_auto_accept_rule_state['reason'],
            response_content
        )

        # Prisoners previous credit
        self.assertIn('£0.10', response_content)
        self.assertIn('01199988199******7253', response_content)
        self.assertIn('02/50', response_content)
        self.assertIn('SOMEONE ELSE', response_content)
        self.assertIn('Number 6', response_content)
        self.assertIn('Soft compliance check failed', response_content)
        self.assertIn(
            prisoner_active_auto_accept_rule_state['reason'],
            response_content
        )
        self.assertNotIn(
            prisoner_inactive_auto_accept_rule_state['reason'],
            response_content
        )

    @parameterized.expand(
        CHECK_REJECTION_CATEGORY_BOOLEAN_MAPPING.items()
    )
    def test_credit_history_row_has_reason_checkbox_populated_for_prisoner_check(
        self, rejection_reason_key, rejection_reason_full
    ):
        """
        Test that the view displays checkboxes associated with a credit.
        """
        check_id = 1
        response_len = 4
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 1,
                    'results': [
                        dict(
                            self.PRISONER_CREDIT,
                            security_check=dict(
                                self.PRISONER_CREDIT['security_check'],
                                rejection_reasons={rejection_reason_key: True}
                            )
                        )
                    ],
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            response = self.client.get(
                reverse(
                    'security:resolve_check',
                    kwargs={'check_id': check_id},
                ),
            )

            self.assertContains(response, rejection_reason_full)

    @parameterized.expand(
        CHECK_REJECTION_CATEGORY_BOOLEAN_MAPPING.items()
    )
    def test_credit_history_row_has_reason_checkbox_populated_for_sender_check(
        self, rejection_reason_key, rejection_reason_full
    ):
        """
        Test that the view displays checkboxes associated with a credit.
        """
        check_id = 1
        response_len = 2
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            sender_credit_response = dict(
                self.SENDER_CREDIT,
                id=self.credit_id + 1,
                security_check=dict(
                    self.SENDER_CREDIT['security_check'],
                    rejection_reasons={rejection_reason_key: True}
                )
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 1,
                    'results': [
                        sender_credit_response
                    ]
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', f'{self.credit_id},{sender_credit_response["id"]}'),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_prisoner_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            response = self.client.get(
                reverse(
                    'security:resolve_check',
                    kwargs={'check_id': check_id},
                ),
            )

            self.assertContains(response, rejection_reason_full)

    @parameterized.expand(
        generate_rejection_category_test_cases_text()
    )
    def test_credit_history_row_has_string_reason_populated_for_prisoner_check(
        self, rejection_reason_key, rejection_reason_value
    ):
        """
        Test that the view displays the reason for associated populated checkboxes with a credit.
        """
        check_id = 1
        response_len = 2
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 1,
                    'results': [
                        dict(
                            self.PRISONER_CREDIT,
                            security_check=dict(
                                self.PRISONER_CREDIT['security_check'],
                                rejection_reasons={rejection_reason_key: rejection_reason_value}
                            )
                        )
                    ]
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            response = self.client.get(
                reverse(
                    'security:resolve_check',
                    kwargs={'check_id': check_id},
                ),
            )

            self.assertContains(response, rejection_reason_value)

    @parameterized.expand(
        generate_rejection_category_test_cases_text()
    )
    def test_credit_history_row_has_string_reason_populated_for_sender_check(
        self, rejection_reason_key, rejection_reason_value
    ):
        """
        Test that the view displays the reason for associated populated checkboxes with a credit.
        """
        check_id = 1
        response_len = 2
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            sender_credit_response = dict(
                self.SENDER_CREDIT,
                id=self.credit_id + 1,
                security_check=dict(
                    self.SENDER_CREDIT['security_check'],
                    rejection_reasons={rejection_reason_key: rejection_reason_value}
                )
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 1,
                    'results': [
                        sender_credit_response,
                    ]
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', f'{self.credit_id},{sender_credit_response["id"]}'),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            response = self.client.get(
                reverse(
                    'security:resolve_check',
                    kwargs={'check_id': check_id},
                ),
            )

            self.assertContains(response, rejection_reason_value)

    def test_accept_check(self):
        """
        Test that if one tries to accept pending check, check marked as accepted

        """
        check_id = 1
        payload_values = {
            'decision_reason': '',
        }
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.POST,
                api_url(f'/security/checks/{check_id}/accept/'),
                match=[
                    responses.json_params_matcher(
                        payload_values
                    )
                ],
                status=204,
            )
            self.mock_need_attention_count(rsps, 0)

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.post(
                url,
                data={
                    'fiu_action': 'accept',
                },
                follow=True
            )

            self.assertRedirects(response, reverse('security:check_list'))
            self.assertContains(response, 'Credit accepted')

    def test_accept_check_without_reactivating_auto_accept(self):
        """
        Test that if one tries to accept pending check, check marked as accepted, without reactivating auto_accept

        We implicitly assert that no API call is made to the security/check/auto-accept endpoint even if
        check.auto_accept_rule populated
        """
        check_id = 1
        auto_accept_rule_id = 35
        payload_values = {
            'decision_reason': '',
        }
        check_with_inactive_auto_accept = copy.deepcopy(self.SENDER_CHECK)
        check_with_inactive_auto_accept['auto_accept_rule_state'] = {
            'auto_accept_rule': auto_accept_rule_id
        }
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=check_with_inactive_auto_accept
            )
            rsps.add(
                rsps.POST,
                api_url(f'/security/checks/{check_id}/accept/'),
                match=[
                    responses.json_params_matcher(
                        payload_values
                    )
                ],
                status=204,
            )
            self.mock_need_attention_count(rsps, 0)

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.post(
                url,
                data={
                    'fiu_action': 'accept',
                },
                follow=True
            )

            self.assertRedirects(response, reverse('security:check_list'))
            self.assertContains(response, 'Credit accepted')

    def test_accept_check_with_auto_accept(self):
        """
        Test that if one tries to accept pending check, check marked as accepted, and auto accept added in active state

        We assert that API call is made to POST security/check/auto-accept endpoint if cleaned_data.auto_accept_reason
        populated
        """
        check_id = 1
        payload_values = {
            'decision_reason': '',
        }
        auto_accept_payload_values = {
            'prisoner_profile': self.SENDER_CHECK['credit']['prisoner_profile'],
            'debit_card_sender_details': self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details'],
            'states': [{
                'reason': 'cause I said so'
            }]
        }
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.POST,
                api_url(f'/security/checks/{check_id}/accept/'),
                match=[
                    responses.json_params_matcher(
                        payload_values
                    )
                ],
                status=204,
            )
            rsps.add(
                rsps.POST,
                api_url('/security/checks/auto-accept'),
                match=[
                    responses.json_params_matcher(
                        auto_accept_payload_values
                    )
                ],
                status=201
            )
            self.mock_need_attention_count(rsps, 0)

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.post(
                url,
                data={
                    'fiu_action': 'accept',
                    'auto_accept_reason': 'cause I said so'
                },
                follow=True
            )

            self.assertRedirects(response, reverse('security:check_list'))
            self.assertContains(response, 'Credit accepted')

    def test_accept_check_with_auto_accept_integrity_error(self):
        """
        Test that if one tries to accept pending check, check marked as accepted, info message displayed

        We assert that API call is made to POST security/check/auto-accept endpoint if cleaned_data.auto_accept_reason
        populated
        """
        check_id = 1
        payload_values = {
            'decision_reason': '',
        }
        auto_accept_payload_values = {
            'prisoner_profile': self.SENDER_CHECK['credit']['prisoner_profile'],
            'debit_card_sender_details': self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details'],
            'states': [{
                'reason': 'cause I said so'
            }]
        }
        check_get_api_url = api_url(f'/security/checks/{check_id}/')
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                check_get_api_url,
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.POST,
                api_url(f'/security/checks/{check_id}/accept/'),
                match=[
                    responses.json_params_matcher(
                        payload_values
                    )
                ],
                status=204,
            )
            rsps.add(
                rsps.POST,
                api_url('/security/checks/auto-accept'),
                match=[
                    responses.json_params_matcher(
                        auto_accept_payload_values
                    )
                ],
                status=400,
                json={
                    'non_field_errors': [
                        CHECK_AUTO_ACCEPT_UNIQUE_CONSTRAINT_ERROR
                    ]
                }
            )
            self.mock_need_attention_count(rsps, 0)

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.post(
                url,
                data={
                    'fiu_action': 'accept',
                    'auto_accept_reason': 'cause I said so'
                },
                follow=True
            )

            self.assertRedirects(response, reverse('security:check_list'))
            self.assertContains(response, 'Credit accepted')
            self.assertContains(
                response,
                (
                    'The auto-accept rule could not be created because an auto-accept rule '
                    'already exists for {sender_name} and {prisoner_number}'.format(
                        sender_name=self.SENDER_CHECK['credit']['sender_name'],
                        prisoner_number=self.SENDER_CHECK['credit']['prisoner_number']
                    )
                )
            )

    def test_accept_check_when_reactivating_auto_accept(self):
        """
        Test that if one tries to accept pending check, check marked as accepted, and auto accept reactivated

        We assert that API call is made to PATCH security/check/auto-accept endpoint if cleaned_data.auto_accept_reason
        and check.auto_accept_rule populated

        """
        check_id = 1
        auto_accept_rule_id = 35
        payload_values = {
            'decision_reason': '',
        }
        check_with_inactive_auto_accept = copy.deepcopy(self.SENDER_CHECK)
        check_with_inactive_auto_accept['auto_accept_rule_state'] = {
            'auto_accept_rule': auto_accept_rule_id
        }
        auto_accept_payload_values = {
            'states': [{
                'active': True,
                'reason': 'cause I said so'
            }]
        }
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=check_with_inactive_auto_accept
            )
            rsps.add(
                rsps.POST,
                api_url(f'/security/checks/{check_id}/accept/'),
                match=[
                    responses.json_params_matcher(
                        payload_values
                    )
                ],
                status=201,
            )
            rsps.add(
                rsps.PATCH,
                api_url(f'/security/checks/auto-accept/{auto_accept_rule_id}'),
                match=[
                    responses.json_params_matcher(
                        auto_accept_payload_values
                    )
                ],
                status=200,
            )
            self.mock_need_attention_count(rsps, 0)

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.post(
                url,
                data={
                    'fiu_action': 'accept',
                    'auto_accept_reason': 'cause I said so'
                },
                follow=True
            )

            self.assertRedirects(response, reverse('security:check_list'))
            self.assertContains(response, 'Credit accepted')

    @parameterized.expand(
        generate_rejection_category_test_cases_text() + generate_rejection_category_test_cases_bool()
    )
    def test_reject_check(self, rejection_field_key, rejection_field_value):
        """
        Test that if a pending check is rejected, the view redirects to the list of checks
        and a successful message is displayed.
        """
        check_id = 1
        payload_values = {
            'decision_reason': 'iamfurtherdetails',
            'rejection_reasons': {rejection_field_key: rejection_field_value}
        }
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.POST,
                api_url(f'/security/checks/{check_id}/reject/'),
                match=[
                    responses.json_params_matcher(
                        payload_values
                    )
                ],
                status=204,
            )
            self.mock_need_attention_count(rsps, 0)

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.post(
                url,
                data={
                    'reject_further_details': 'iamfurtherdetails',
                    rejection_field_key: rejection_field_value,
                    'fiu_action': 'reject',
                },
                follow=True,
            )

            self.assertRedirects(response, reverse('security:check_list'))
            self.assertContains(response, 'Credit rejected')

    def test_invalid_if_check_not_in_pending(self):
        """
        Test that if one tries to reject an already accepted check, a validation error is displayed.
        """
        check_id = 1
        response_len = 4
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK_REJECTED
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len))
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_prisoner_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.post(
                url,
                data={
                    'payment_source_paying_multiple_prisoners': True,
                    'fiu_action': 'reject',
                },
                follow=True,
            )

            self.assertContains(response, 'You cannot action this credit')

    def test_invalid_with_no_rejection_reason(self):
        """
        Test that if the rejection reason is not given, a validation error is displayed.
        """
        check_id = 1
        response_len = 2
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_sender_credit_list(response_len)),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': response_len,
                    'results': list(self._get_prisoner_credit_list(response_len))
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})
            response = self.client.post(
                url,
                data={
                    'fiu_action': 'reject',
                },
                follow=True,
            )

            self.assertContains(response, 'You must provide a reason for rejecting a credit')

    def test_credit_history_ordering(self):
        """
        Test that the credit history is correctly ordered by `started_at`
        """
        sender_credits = [
            dict(
                self.SENDER_CREDIT,
                **{
                    'id': i,
                    'security_check': dict(
                        self.SENDER_CREDIT['security_check'],
                        actioned_at=offset_isodatetime_by_ten_seconds(
                            self.SENDER_CREDIT['security_check']['actioned_at'], i
                        )
                    )
                }
            )
            for i in range(4)
        ]
        prisoner_credits = [
            dict(
                self.PRISONER_CREDIT,
                **{
                    'id': i,
                    'security_check': dict(
                        self.PRISONER_CREDIT['security_check'],
                        actioned_at=offset_isodatetime_by_ten_seconds(
                            self.PRISONER_CREDIT['security_check']['actioned_at'], i
                        )
                    )
                }
            )
            for i in range(4)
        ]
        # We expect the credits to be ordered according to started at.
        # Given the offsets above we expect the first sender credit followed by the first prisoner credit
        # and so on
        expected_related_credits = list(itertools.chain(*zip(sender_credits, prisoner_credits)))
        # We want to display the newest credits first
        expected_related_credits.reverse()
        mock_api_session = mock.MagicMock()
        mock_api_session.configure_mock(**{
            'get.return_value.json.return_value.get': mock.MagicMock(
                side_effect=[
                    4,
                    sender_credits,
                    4,
                    prisoner_credits
                ]
            )
        })
        actual_related_credits = AcceptOrRejectCheckView()._get_related_credits(
            api_session=mock_api_session,
            detail_object={
                'credit': {
                    'id': self.credit_id,
                    'prisoner_profile': self.prisoner_id,
                    'sender_profile': self.sender_id
                }
            }
        )
        self.assertListEqual(expected_related_credits, actual_related_credits)


class CheckAssignViewTestCase(BaseCheckViewTestCase, SecurityViewTestCase):
    """
    Tests related to CheckAssignView.
    """
    sender_id = 2
    prisoner_id = 3
    credit_id = 5

    SENDER_CHECK = copy.deepcopy(BaseCheckViewTestCase.SAMPLE_CHECK)
    SENDER_CHECK['credit']['sender_profile'] = sender_id
    SENDER_CHECK['credit']['prisoner_profile'] = prisoner_id
    SENDER_CHECK['credit']['id'] = credit_id

    def test_assign_view_get_request_redirects_to_resolve_check_page(self):
        """
        A GET request should redirect to the resolve page
        For instance if a session expires and the user hits 'Add to my list'
        """
        check_id = 1
        response_len = 0
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=dict(self.SENDER_CHECK, assigned_to_name='Columbo', assigned_to=777)
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:assign_check', kwargs={'check_id': check_id})
            response = self.client.get(url, follow=True)

            self.assertRedirects(response, reverse('security:resolve_check', kwargs={'check_id': check_id}))

    def test_assign_view_get_request_redirects_to_resolve_check_page_with_kwarg(self):
        """
        A GET request should redirect to the resolve page
        For instance if a session expires and the user hits 'Add to my list'
        """
        check_id = 1
        response_len = 0
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=dict(self.SENDER_CHECK, assigned_to_name='Columbo', assigned_to=777)
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:assign_check', kwargs={'check_id': check_id})

            response = self.client.get(url, follow=True)

            self.assertRedirects(response, reverse('security:resolve_check', kwargs={'check_id': check_id}))

    def test_assign_view_get_request_redirects_to_list_check_page_with_kwarg(self):
        """
        A GET request should redirect to the list page if called with 'list' as positional url arg
        """
        check_id = 1
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 1,
                    'data': [self.SAMPLE_CHECK]
                }
            )

            url = reverse('security:assign_check', kwargs={'check_id': check_id, 'current_page': 1, 'list': 'list'})

            response = self.client.get(url, follow=True)

            self.assertRedirects(
                response, reverse('security:check_list') + f'?page=1#check-row-{check_id}'
            )

    def test_assign_view_get_request_redirects_to_list_check_page_with_kwarg_and_current_page(self):
        """
        A GET request should redirect to the list page if called with 'list' as positional url arg
        """
        check_id = 1

        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url('/security/checks/'),
                json={
                    'count': 21,
                    'data': [self.SAMPLE_CHECK] * 21
                }
            )

            url = reverse('security:assign_check', kwargs={'check_id': check_id, 'current_page': 2, 'list': 'list'})

            response = self.client.get(url, follow=True)

            self.assertRedirects(
                response, reverse('security:check_list') + f'?page=2#check-row-{check_id}'
            )
            self.assertContains(response, 'Page 2 of 2')

    def test_can_assign_check_to_own_list(self):
        """
        Test that a user can add a check to their own list of checks
        """
        check_id = 1
        response_len = 0
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.PATCH,
                api_url(f'/security/checks/{check_id}/'),
                json={
                    'count': 1,
                    'results': dict(self.SENDER_CHECK, assigned_to_name='Sherlock Holmes', assigned_to=5),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=dict(self.SENDER_CHECK, assigned_to_name='Sherlock Holmes', assigned_to=5)
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:assign_check', kwargs={'check_id': check_id})

            response = self.client.post(
                url,
                data={
                    'assignment': 'assign',
                },
                follow=True,
            )

            self.assertRedirects(response, reverse('security:resolve_check', kwargs={'check_id': check_id}))
            self.assertContains(response, 'Remove from my list')

    def test_can_remove_check_from_their_list(self):
        """
        Test that a user can add a check to their own list of checks
        """
        check_id = 1
        response_len = 0
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.PATCH,
                api_url(f'/security/checks/{check_id}/'),
                json={
                    'count': 1,
                    'results': dict(self.SENDER_CHECK, assigned_to_name=None, assigned_to=None),
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=dict(self.SENDER_CHECK, assigned_to_name=None, assigned_to=None)
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:assign_check', kwargs={'check_id': check_id})

            response = self.client.post(
                url,
                data={
                    'assignment': 'unassign',
                },
                follow=True,
            )

            calls = list(rsps.calls)
            self.assertTrue(json.loads(calls[3].request.body) == {'assigned_to': None})
            self.assertRedirects(response, reverse('security:resolve_check', kwargs={'check_id': check_id}))
            self.assertContains(response, 'Add to my list')

    def test_resolve_page_displays_other_username_if_assigned_else(self):
        """
        Test that a user can see that a different user has already assigned the check to their list
        """
        check_id = 1
        response_len = 0
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=dict(self.SENDER_CHECK, assigned_to_name='Joe Bloggs', assigned_to=200)
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:resolve_check', kwargs={'check_id': check_id})

            response = self.client.get(
                url,
                follow=True,
            )

            self.assertNotContains(response, 'name="assignment"')
            self.assertContains(response, 'Added to Joe Bloggs’ list')

    @mock.patch('security.forms.check.get_need_attention_date')
    def test_resolve_page_displays_error_if_assignment_collision_from_list(self, mock_get_need_attention_date):
        """
        Test that a user sees an error if trying to assign check to self that's already assigned to someone else from
        check list view
        """
        mock_get_need_attention_date.return_value = make_aware(datetime.datetime(2019, 7, 2, 9))
        check_id = 1
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.PATCH,
                api_url(f'/security/checks/{check_id}/'),
                status=400,
                json=[
                    'That check is already assigned to Someone Else'
                ]
            )
            self.mock_need_attention_count(rsps, 1)

            url = reverse('security:assign_check', kwargs={'check_id': check_id, 'list': 'list'})

            with silence_logger():
                response = self.client.post(
                    url,
                    follow=True,
                    data={'assignment': 'assign'}
                )

            self.assertNotContains(response, 'name="assignment"')
            self.assertContains(response, 'That check is already assigned to Someone Else')
            self.assertRedirects(response, reverse('security:check_list'))

    def test_resolve_page_displays_error_if_assignment_collision_from_accept_reject(self):
        """
        Test that a user sees an error if trying to assign check to self that's already assigned to someone else from
        accept reject check view
        """
        check_id = 1
        response_len = 0
        with responses.RequestsMock() as rsps:
            self.login(rsps)
            rsps.add(
                rsps.GET,
                api_url(f'/security/checks/{check_id}/'),
                json=self.SENDER_CHECK
            )
            rsps.add(
                rsps.PATCH,
                api_url(f'/security/checks/{check_id}/'),
                status=400,
                json=[
                    'That check is already assigned to Someone Else'
                ]
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/senders/{sender_profile_id}/credits/?{querystring}'.format(
                        sender_profile_id=self.sender_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', self.credit_id),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                urljoin(
                    settings.API_URL,
                    '/prisoners/{prisoner_profile_id}/credits/?{querystring}'.format(
                        prisoner_profile_id=self.prisoner_id,
                        querystring=urlencode([
                            ('limit', 500),
                            ('offset', 0),
                            ('exclude_credit__in', ','.join(map(str, ([self.credit_id] + list(range(response_len)))))),
                            ('security_check__isnull', False),
                            ('only_completed', False),
                            ('security_check__actioned_by__isnull', False),
                            ('include_checks', True)
                        ])
                    ),
                    trailing_slash=False
                ),
                json={
                    'count': 0,
                }
            )
            rsps.add(
                rsps.GET,
                '{}/security/checks/auto-accept/?{}'.format(
                    settings.API_URL,
                    urlencode((
                        ('prisoner_profile_id', self.SENDER_CHECK['credit']['prisoner_profile']),
                        (
                            'debit_card_sender_details_id',
                            self.SENDER_CHECK['credit']['billing_address']['debit_card_sender_details']
                        )
                    ))
                ),
                json={'count': 0, 'prev': None, 'next': None, 'results': []}
            )

            url = reverse('security:assign_check', kwargs={'check_id': check_id})

            with silence_logger():
                response = self.client.post(
                    url,
                    follow=True,
                    data={'assignment': 'assign'}
                )

            self.assertNotContains(response, 'name="assignment"')
            self.assertContains(response, 'That check is already assigned to Someone Else')
            self.assertRedirects(response, reverse('security:resolve_check', kwargs={'check_id': check_id}))


class PolicyChangeViewTestCase(SecurityBaseTestCase):
    @responses.activate
    @override_settings(NOVEMBER_SECOND_CHANGES_LIVE=False)
    def test_displays_policy_warning_page_before_policy_change(self):
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
        response = self.client.get(reverse('security:policy_change'), follow=True)
        self.assertContains(response, 'policy-change-warning')

    @responses.activate
    @override_settings(NOVEMBER_SECOND_CHANGES_LIVE=True)
    def test_displays_policy_update_page_after_policy_change(self):
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
        response = self.client.get(reverse('security:policy_change'), follow=True)
        self.assertContains(response, 'policy-change-info')
