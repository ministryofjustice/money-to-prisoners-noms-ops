from contextlib import contextmanager
import tempfile
from unittest import mock

from django.conf import settings
from django.test import SimpleTestCase, override_settings
from django.urls import reverse
from mtp_common.auth import urljoin
from mtp_common.auth.api_client import get_request_token_url
from mtp_common.auth.test_utils import generate_tokens
from openpyxl import load_workbook

from security import (
    confirmed_prisons_flag,
    hmpps_employee_flag,
    required_permissions,
    provided_job_info_flag,
)
from security.tests import api_url

SAMPLE_PRISONS = [
    {
        'nomis_id': 'AAI',
        'general_ledger_code': '001',
        'name': 'HMP & YOI Test 1',
        'short_name': 'Test 1',
        'region': 'London',
        'categories': [{'description': 'Category D', 'name': 'D'},
                       {'description': 'Young Offender Institution', 'name': 'YOI'}],
        'populations': [{'description': 'Female', 'name': 'female'},
                        {'description': 'Male', 'name': 'male'},
                        {'description': 'Young offenders', 'name': 'young'}],
        'pre_approval_required': False,
    },
    {
        'nomis_id': 'BBI',
        'general_ledger_code': '002',
        'name': 'HMP Test 2',
        'short_name': 'Test 2',
        'region': 'London',
        'categories': [{'description': 'Category D', 'name': 'D'}],
        'populations': [{'description': 'Male', 'name': 'male'}],
        'pre_approval_required': False,
    },
]


def mock_prison_response(rsps=None):
    import responses
    rsps = rsps or responses
    rsps.add(
        rsps.GET,
        api_url('/prisons/'),
        json={
            'count': len(SAMPLE_PRISONS),
            'results': SAMPLE_PRISONS,
        }
    )


def no_saved_searches(rsps=None):
    import responses
    rsps = rsps or responses
    rsps.add(
        rsps.GET,
        api_url('/searches/'),
        json={
            'count': 0,
            'results': []
        },
    )


@contextmanager
def temp_spreadsheet(data):
    with tempfile.TemporaryFile() as f:
        f.write(data)
        wb = load_workbook(f)
        ws = wb.active
        yield ws


@override_settings(
    CACHES={'default': {'BACKEND': 'django.core.cache.backends.dummy.DummyCache'}},
)
class SecurityBaseTestCase(SimpleTestCase):
    mock_user_pk = 5

    def setUp(self):
        super().setUp()
        self.notifications_mock = mock.patch('mtp_common.templatetags.mtp_common.notifications_for_request',
                                             return_value=[])
        self.notifications_mock.start()

    def tearDown(self):
        self.notifications_mock.stop()
        super().tearDown()

    def login(self, rsps, follow=True, user_data=None):
        no_saved_searches(rsps=rsps)
        return self._login(rsps, follow=follow, user_data=user_data)

    def _login(self, rsps, follow=True, user_data=None):
        returned_user_data = user_data or self.get_user_data()
        returned_user_data['pk'] = self.mock_user_pk
        rsps.add(
            rsps.POST,
            get_request_token_url(),
            json=generate_tokens()
        )
        rsps.add(
            rsps.GET,
            urljoin(settings.API_URL, '/users/{username}'.format(username=returned_user_data['username'])),
            json=returned_user_data
        )

        response = self.client.post(
            reverse('login'),
            data={'username': 'shall', 'password': 'pass'},
            follow=follow
        )
        if follow:
            self.assertEqual(response.status_code, 200)
        else:
            self.assertEqual(response.status_code, 302)
        return response

    def get_user_data(
        self,
        first_name='Sam',
        last_name='Hall',
        username='shall',
        email='sam@mtp.local',
        permissions=required_permissions,
        prisons=(SAMPLE_PRISONS[1],),
        flags=(hmpps_employee_flag, confirmed_prisons_flag, provided_job_info_flag),
        roles=('security',),
    ):
        return {
            'first_name': first_name,
            'last_name': last_name,
            'username': username,
            'email': email,
            'permissions': permissions,
            'prisons': prisons,
            'flags': flags,
            'roles': roles,
        }

    def assertSpreadsheetEqual(self, spreadsheet_data, expected_values, msg=None):  # noqa: N802
        with temp_spreadsheet(spreadsheet_data) as ws:
            for i, row in enumerate(expected_values, start=1):
                for j, cell in enumerate(row, start=1):
                    self.assertEqual(cell, ws.cell(column=j, row=i).value, msg=msg)
