from contextlib import contextmanager
from unittest import mock
import tempfile

from django.core import mail
from django.core.urlresolvers import reverse
from django.test import SimpleTestCase, override_settings
from mtp_common.auth.test_utils import generate_tokens
from openpyxl import load_workbook
import responses

from security import (
    confirmed_prisons_flag,
    hmpps_employee_flag,
    required_permissions,
    SEARCH_V2_FLAG,
)
from security.tests.utils import api_url
from security.views.base import SIMPLE_SEARCH_FORM_SUBMITTED_INPUT_NAME


@contextmanager
def temp_spreadsheet(data):
    with tempfile.TemporaryFile() as f:
        f.write(data)
        wb = load_workbook(f)
        ws = wb.active
        yield ws


override_nomis_settings = override_settings(
    NOMIS_API_BASE_URL='https://nomis.local/',
    NOMIS_API_CLIENT_TOKEN='hello',
    NOMIS_API_PRIVATE_KEY=(
        '-----BEGIN EC PRIVATE KEY-----\n'
        'MHcCAQEEIOhhs3RXk8dU/YQE3j2s6u97mNxAM9s+13S+cF9YVgluoAoGCCqGSM49\n'
        'AwEHoUQDQgAE6l49nl7NN6k6lJBfGPf4QMeHNuER/o+fLlt8mCR5P7LXBfMG6Uj6\n'
        'TUeoge9H2N/cCafyhCKdFRdQF9lYB2jB+A==\n'
        '-----END EC PRIVATE KEY-----\n'
    ),  # this key is just for tests, doesn't do anything
)

sample_prisons = [
    {
        'nomis_id': 'AAI',
        'general_ledger_code': '001',
        'name': 'HMP & YOI Test 1', 'short_name': 'Test 1',
        'region': 'London',
        'categories': [{'description': 'Category D', 'name': 'D'},
                       {'description': 'Young Offender Institution', 'name': 'YOI'}],
        'populations': [{'description': 'Female', 'name': 'female'},
                        {'description': 'Male', 'name': 'male'},
                        {'description': 'Young offenders', 'name': 'young'}],
        'pre_approval_required': False,
    },
    {
        'nomis_id': 'BBI',
        'general_ledger_code': '002',
        'name': 'HMP Test 2', 'short_name': 'Test 2',
        'region': 'London',
        'categories': [{'description': 'Category D', 'name': 'D'}],
        'populations': [{'description': 'Male', 'name': 'male'}],
        'pre_approval_required': False,
    },
]
default_user_prisons = (sample_prisons[1],)


def sample_prison_list(rsps=None):
    rsps = rsps or responses
    rsps.add(
        rsps.GET,
        api_url('/prisons/'),
        json={
            'count': len(sample_prisons),
            'results': sample_prisons,
        }
    )


def no_saved_searches(rsps=None):
    rsps = rsps or responses
    rsps.add(
        rsps.GET,
        api_url('/searches/'),
        json={
            'count': 0,
            'results': []
        },
    )


@override_settings(
    CACHES={'default': {'BACKEND': 'django.core.cache.backends.dummy.DummyCache'}},
)
class SecurityBaseTestCase(SimpleTestCase):

    def setUp(self):
        super().setUp()
        self.notifications_mock = mock.patch('mtp_common.templatetags.mtp_common.notifications_for_request',
                                             return_value=[])
        self.notifications_mock.start()

    def tearDown(self):
        self.notifications_mock.stop()
        super().tearDown()

    @mock.patch('mtp_common.auth.backends.api_client')
    def login(self, mock_api_client, follow=True, user_data=None, rsps=None):
        no_saved_searches(rsps=rsps)
        return self._login(mock_api_client, follow=follow, user_data=user_data)

    @mock.patch('mtp_common.auth.backends.api_client')
    def login_test_searches(self, mock_api_client, follow=True):
        return self._login(mock_api_client, follow=follow)

    def _login(self, mock_api_client, follow=True, user_data=None):
        mock_api_client.authenticate.return_value = {
            'pk': 5,
            'token': generate_tokens(),
            'user_data': user_data or self.get_user_data()
        }

        response = self.client.post(
            reverse('login'),
            data={'username': 'shall', 'password': 'pass'},
            follow=follow
        )
        if follow:
            self.assertEqual(response.status_code, 200)
        else:
            self.assertEqual(response.status_code, 302)
        return response

    def get_user_data(
        self, first_name='Sam', last_name='Hall', username='shall',
        email='sam@mtp.local', permissions=required_permissions,
        prisons=default_user_prisons,
        flags=(hmpps_employee_flag, confirmed_prisons_flag,),
        roles=('security',)
    ):
        return {
            'first_name': first_name,
            'last_name': last_name,
            'username': username,
            'email': email,
            'permissions': permissions,
            'prisons': prisons,
            'flags': flags,
            'roles': roles,
        }

    def assertSpreadsheetEqual(self, spreadsheet_data, expected_values, msg=None):  # noqa: N802
        with temp_spreadsheet(spreadsheet_data) as ws:
            for i, row in enumerate(expected_values, start=1):
                for j, cell in enumerate(row, start=1):
                    self.assertEqual(cell, ws.cell(column=j, row=i).value, msg=msg)


class SecurityViewTestCase(SecurityBaseTestCase):
    view_name = None
    api_list_path = None
    search_ordering = None

    bank_transfer_sender = {
        'id': 9,
        'credit_count': 4,
        'credit_total': 41000,
        'prisoner_count': 3,
        'prison_count': 2,
        'bank_transfer_details': [
            {
                'sender_name': 'MAISIE NOLAN',
                'sender_sort_code': '101010',
                'sender_account_number': '12312345',
                'sender_roll_number': '',
            }
        ],
        'created': '2016-05-25T20:24:00Z',
        'modified': '2016-05-25T20:24:00Z',
    }
    debit_card_sender = {
        'id': 9,
        'credit_count': 4,
        'credit_total': 42000,
        'prisoner_count': 3,
        'prison_count': 2,
        'bank_transfer_details': [],
        'debit_card_details': [
            {
                'card_number_last_digits': '1234',
                'card_expiry_date': '10/20',
                'postcode': 'SW137NJ',
                'sender_emails': ['m@outside.local', 'M@OUTSIDE.LOCAL', 'mn@outside.local'],
                'cardholder_names': ['Maisie N', 'MAISIE N', 'Maisie Nolan'],
            }
        ],
        'created': '2016-05-25T20:24:00Z',
        'modified': '2016-05-25T20:24:00Z',
    }
    prisoner_profile = {
        'id': 8,
        'credit_count': 3,
        'credit_total': 31000,
        'sender_count': 2,
        'disbursement_count': 2,
        'disbursement_total': 29000,
        'recipient_count': 1,
        'prisoner_name': 'JAMES HALLS',
        'prisoner_number': 'A1409AE',
        'prisoner_dob': '1986-12-09',
        'current_prison': {'nomis_id': 'PRN', 'name': 'Prison'},
        'prisons': [{'nomis_id': 'PRN', 'name': 'Prison'}],
        'provided_names': ['Jim Halls', 'JAMES HALLS', 'James Halls '],
        'created': '2016-05-25T20:24:00Z',
        'modified': '2016-05-25T20:24:00Z',
    }
    credit_object = {
        'id': 1,
        'amount': 10250,
        'resolution': 'credited', 'anonymous': False,
        'credited_at': '2016-05-25T20:24:00Z',
        'owner': 1, 'owner_name': 'Clerk',
        'prison': 'PRN', 'prison_name': 'Prison', 'prisoner_name': 'JAMES HALLS', 'prisoner_number': 'A1409AE',
        'reconciliation_code': None,
        'received_at': '2017-01-25T12:00:00Z',
        'refunded_at': None,
        'reviewed': False, 'comments': [],
        'source': 'bank_transfer',
        'sender_sort_code': '101010', 'sender_account_number': '12312345', 'sender_roll_number': '',
        'card_expiry_date': None, 'card_number_last_digits': None,
        'sender_name': 'MAISIE NOLAN',
        'sender_email': None,
        'intended_recipient': None,
    }

    def get_api_object_list_response_data(self):
        """
        Return the list of objects that the mocked api should return in the tests.
        """
        return []

    @responses.activate
    @mock.patch('security.forms.base.SecurityForm.get_object_list')
    def test_can_access_security_view(self, mocked_form_method):
        mocked_form_method.return_value = []
        if not self.view_name:
            return
        sample_prison_list()
        self.login()
        response = self.client.get(reverse(self.view_name), follow=True)
        self.assertContains(response, '<!-- %s -->' % self.view_name)

    @responses.activate
    def test_filtering_by_one_prison(self):
        if not self.api_list_path:
            return
        self.login()
        no_saved_searches()
        sample_prison_list()
        responses.add(responses.GET, api_url(self.api_list_path), json={'count': 0, 'results': []})
        self.client.get(reverse(self.view_name) + '?page=1&prison=BBI', follow=False)
        calls = list(filter(lambda call: self.api_list_path in call.request.url, responses.calls))
        self.assertEqual(len(calls), 1)
        self.assertIn('prison=BBI', calls[0].request.url)

    @responses.activate
    def test_filtering_by_many_prisons(self):
        if not self.api_list_path:
            return
        self.login()
        no_saved_searches()
        sample_prison_list()
        responses.add(responses.GET, api_url(self.api_list_path), json={'count': 0, 'results': []})
        self.client.get(reverse(self.view_name) + '?page=1&prison=BBI&prison=AAI', follow=False)
        calls = list(filter(lambda call: self.api_list_path in call.request.url, responses.calls))
        self.assertEqual(len(calls), 1)
        self.assertIn('prison=AAI&prison=BBI', calls[0].request.url)

    @responses.activate
    def test_filtering_by_many_prisons_alternate(self):
        if not self.api_list_path:
            return
        self.login()
        no_saved_searches()
        sample_prison_list()
        responses.add(responses.GET, api_url(self.api_list_path), json={'count': 0, 'results': []})
        self.client.get(reverse(self.view_name) + '?page=1&prison=BBI,AAI', follow=False)
        calls = list(filter(lambda call: self.api_list_path in call.request.url, responses.calls))
        self.assertEqual(len(calls), 1)
        self.assertIn('prison=AAI&prison=BBI', calls[0].request.url)


class SimpleSearchV2SecurityTestCaseMixin:
    search_results_view_name = None

    def get_user_data(
        self,
        *args,
        flags=(
            hmpps_employee_flag,
            confirmed_prisons_flag,
            SEARCH_V2_FLAG,
        ),
        **kwargs,
    ):
        """
        Sets the SEARCH_V2_FLAG feature flag by default.
        """

        return super().get_user_data(*args, flags=flags, **kwargs)

    def test_simple_search_displays_search_results(self):
        """
        Test that the search results page includes the objects returned by the API.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 2,
                    'results': self.get_api_object_list_response_data(),
                },
            )
            response = self.client.get(reverse(self.view_name))
        self._test_simple_search_search_results_content(response)

    def _test_simple_search_search_results_content(self, response):
        """
        Subclass to test that the response content of the search results view is as expected.
        """
        raise NotImplementedError

    def test_simple_search_uses_default_prisons(self):
        """
        Test that the simple search template uses the prisons of the logged in user to
        populate the `prison` hidden inputs.
        The simple search filters by the user's prisons by default.
        """
        with responses.RequestsMock() as rsps:
            user_prisons = sample_prisons
            user_data = self.get_user_data(prisons=user_prisons)

            self.login(user_data=user_data, rsps=rsps)
            sample_prison_list(rsps=rsps)
            response = self.client.get(reverse(self.view_name))

        for prison in user_prisons:
            self.assertContains(
                response,
                f'<input type="hidden" name="prison" value="{prison["nomis_id"]}" />',
            )

    def test_simple_search_redirects_to_search_results_page(self):
        """
        Test that submitting the form redirects to the results page when the form is valid.
        The action of submitting the form is represented by the query param
        SIMPLE_SEARCH_FORM_SUBMITTED_INPUT_NAME which gets removed when redirecting.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )
            query_string = f'ordering={self.search_ordering}&simple_search=test'
            request_url = f'{reverse(self.view_name)}?{query_string}&{SIMPLE_SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            expected_redirect_url = f'{reverse(self.search_results_view_name)}?{query_string}'
            response = self.client.get(request_url)
            self.assertRedirects(
                response,
                expected_redirect_url,
            )

    def test_doesnt_redirect_to_search_results_page_if_form_is_invalid(self):
        """
        Test that submitting the form doesn't redirect to the results page when the form is invalid.
        The action of submitting the form is represented by the query param
        SIMPLE_SEARCH_FORM_SUBMITTED_INPUT_NAME.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)
            query_string = 'ordering=invalid&simple_search=test'
            request_url = f'{reverse(self.view_name)}?{query_string}&{SIMPLE_SEARCH_FORM_SUBMITTED_INPUT_NAME}=1'
            response = self.client.get(request_url)
            self.assertEqual(response.status_code, 200)

    def test_navigating_back_to_simple_search_form(self):
        """
        Test that going back to the simple search form doesn't redirect to the results page.
        The action of NOT submitting the form explicitly is represented by the absence of query param
        SIMPLE_SEARCH_FORM_SUBMITTED_INPUT_NAME.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )
            query_string = f'ordering={self.search_ordering}&simple_search=test'
            request_url = f'{reverse(self.view_name)}?{query_string}'
            response = self.client.get(request_url)
            self.assertEqual(response.status_code, 200)


class ExportSecurityViewTestCaseMixin:
    export_view_name = None
    export_email_view_name = None
    export_expected_xls_headers = None
    export_expected_xls_rows = None

    def test_export_some_data(self):
        """
        Test that the export view generates a spreadsheet with only the content returned
        by the API.
        """
        expected_spreadsheet_content = [
            self.export_expected_xls_headers,
            *self.export_expected_xls_rows,
        ]

        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 2,
                    'results': self.get_api_object_list_response_data(),
                }
            )
            response = self.client.get(reverse(self.export_view_name))

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertSpreadsheetEqual(response.content, expected_spreadsheet_content)

    def test_email_export_some_data(self):
        """
        Test that the view sends the expected spreadsheet via email.
        """
        expected_spreadsheet_content = [
            self.export_expected_xls_headers,
            *self.export_expected_xls_rows,
        ]

        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)
            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 2,
                    'results': self.get_api_object_list_response_data(),
                }
            )
            response = self.client.get(f'{reverse(self.export_email_view_name)}?ordering={self.search_ordering}')
            self.assertRedirects(response, f'{reverse(self.view_name)}?ordering={self.search_ordering}')
            self.assertSpreadsheetEqual(
                mail.outbox[0].attachments[0][1],
                expected_spreadsheet_content,
                msg='Emailed contents do not match expected',
            )

    def test_export_no_data(self):
        """
        Test that the export view generates a spreadsheet without rows if the API call doesn't return any record.
        """
        expected_spreadsheet_content = [
            self.export_expected_xls_headers,
        ]

        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)

            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )

            response = self.client.get(reverse(self.export_view_name))

        self.assertEqual(200, response.status_code)
        self.assertEqual(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        self.assertSpreadsheetEqual(response.content, expected_spreadsheet_content)

    def test_email_export_no_data(self):
        """
        Test that the view sends the an empty spreadsheet via email.
        """
        expected_spreadsheet_content = [
            self.export_expected_xls_headers,
        ]

        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)

            rsps.add(
                rsps.GET,
                api_url(self.api_list_path),
                json={
                    'count': 0,
                    'results': [],
                },
            )

            response = self.client.get(f'{reverse(self.export_email_view_name)}?ordering={self.search_ordering}')
            self.assertRedirects(response, f'{reverse(self.view_name)}?ordering={self.search_ordering}')
            self.assertSpreadsheetEqual(
                mail.outbox[0].attachments[0][1],
                expected_spreadsheet_content,
                msg='Emailed contents do not match expected',
            )

    def test_invalid_params_redirect_to_form(self):
        """
        Test that in case of invalid form, the export view redirects to the list page without
        taking any action.
        """
        with responses.RequestsMock() as rsps:
            self.login(rsps=rsps)
            sample_prison_list(rsps=rsps)
            response = self.client.get(f'{reverse(self.export_email_view_name)}?ordering=invalid')
            self.assertRedirects(response, f'{reverse(self.view_name)}?ordering=invalid')
